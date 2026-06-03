from django.shortcuts import render, HttpResponseRedirect
from .models import Penjualan, PenjualanDetail
from stock.models import Barang, Cabang, DaftarPaket
from django.contrib.auth.models import User
from django.db.models import Q, Count
from django.db import transaction
from django.contrib import messages
from django.conf import settings
import datetime
from django.contrib.auth import authenticate,login,logout
from django.contrib.sessions.models import Session
from django.utils import timezone
from cms.views import addLog
from promo.models import Promo
from cara.models import Tutorial,TutorialComment,TutorialImage
from cms.models import Testimoni, ResetPasswordToken
from posmimail import posmiMail
import uuid as uuid_module

DAFTAR_PAKET = []
PAKET_GRATIS_KUOTA_BULANAN = 75


def _reset_is_checked(nota):
    """Reset status cek transaksi saat item berubah."""
    try:
        Penjualan.objects.filter(nota=nota, is_paid=False).update(is_checked=False)
    except Exception:
        pass


def _hapus_session_lain(user, current_session_key):
    """Hapus semua session aktif milik user kecuali session saat ini."""
    for session in Session.objects.filter(expire_date__gte=timezone.now()):
        data = session.get_decoded()
        if str(data.get('_auth_user_id')) == str(user.pk) and session.session_key != current_session_key:
            session.delete()


def _hapus_transaksi_pending_user(user):
    """Hapus transaksi pending milik kasir yang sedang logout dan kembalikan stoknya."""
    if not user or not user.is_authenticated:
        return 0

    cabang_id = getattr(getattr(user, 'userprofile', None), 'cabang_id', None)
    total_dihapus = 0
    with transaction.atomic():
        pending_qs = Penjualan.objects.filter(user_id=user.id, is_paid=False)
        if cabang_id:
            pending_qs = pending_qs.filter(cabang_id=cabang_id)
        for penjualan in pending_qs:
            for detail in PenjualanDetail.objects.select_related('barang').filter(penjualan=penjualan):
                barang = detail.barang
                barang.stok += detail.jumlah
                barang.save(update_fields=['stok'])
            penjualan.delete()
            total_dihapus += 1
    return total_dihapus

def _korporasi_aktif(cabang):
    """True jika toko korporasi dan lisensi owner masih berlaku."""
    return cabang.owner_id and (cabang.owner.is_active or cabang.owner.in_grace)


def cek_expired_kuota(id_cabang):
    cabang = Cabang.objects.get(id=id_cabang)
    if cabang.owner_id:
        if _korporasi_aktif(cabang):
            return cabang.owner.kuota_transaksi_pool > 0
        # Korporasi unlicensed → fallback free tier (kuota_transaksi per cabang)
        return cabang.kuota_transaksi > 0
    if cabang.paket is None:
        return cabang.kuota_transaksi > 0
    try:
        if cabang.lisensi_expired < datetime.datetime.now():
            if cabang.lisensi_grace < datetime.datetime.now():
                return False
            return cabang.kuota_transaksi > 0
        return cabang.kuota_transaksi > 0
    except:
        return False

def cek_kuota_transaksi(id_cabang):
    cabang = Cabang.objects.get(id=id_cabang)
    if cabang.owner_id:
        if _korporasi_aktif(cabang):
            return cabang.owner.kuota_transaksi_pool > 0
        return cabang.kuota_transaksi > 0
    return cabang.kuota_transaksi > 0

def cek_lisensi_expired(id_cabang):
    try:
        cabang = Cabang.objects.get(id=id_cabang)
        if cabang.owner_id:
            return _korporasi_aktif(cabang)
        if cabang.paket is None:
            return True
        return cabang.lisensi_expired >= datetime.datetime.now()
    except:
        return False

def cek_jumlah_grace(id_cabang):
    try:
        cabang = Cabang.objects.get(id=id_cabang)
        if cabang.owner_id:
            return cabang.owner.sisa_hari_grace
        return (cabang.lisensi_grace - datetime.datetime.now()).days
    except:
        return 0

def index(request):
    toko = None
    tanggal = datetime.datetime.now()
    cabang_available=False
    jumlah_grace = 0
    jumlah_transaksi=0
    bisa_transaksi=False
    tutorial = Tutorial.objects.annotate(
        comment_count=Count('tutorialcomment')
    )[:3]
    # Buat dict image per tutorial — 1 query
    tutorialimage = TutorialImage.objects.filter(tutorial__in=tutorial)
    image_map = {ti.tutorial_id: ti.image for ti in tutorialimage}
    image = []
    for tutor in tutorial:
        img = image_map.get(tutor.id)
        if img:
            image.append({
                'tutorial': tutor,
                'image': img,
                'comment': tutor.comment_count,
            })
    
    try:
        limapenjualan = Penjualan.objects.all().filter(Q(user=request.user) & Q(is_paid=True)).order_by('-tgl_bayar')[:5]
    except:
        limapenjualan = None

    if request.user.is_authenticated:
        if request.user.is_staff:
            return HttpResponseRedirect('/management/')
        # Owner Korporasi tidak punya userprofile → arahkan ke dashboard owner
        if hasattr(request.user, 'owner_profile'):
            return HttpResponseRedirect('/owner/')
        jumlah_transaksi = request.user.userprofile.cabang.kuota_transaksi
        cabang_obj = request.user.userprofile.cabang
        is_paket_gratis = cabang_obj.paket is None and cabang_obj.owner_id is None
        is_paket_korporasi = cabang_obj.owner_id is not None
        pool_kuota_korporasi = cabang_obj.owner.kuota_transaksi_pool if is_paket_korporasi else 0
        owner_session = request.session.get('owner_user_id') is not None
        korporasi_unlicensed = is_paket_korporasi and not _korporasi_aktif(cabang_obj)
        cabang_available = cek_expired_kuota(request.user.userprofile.cabang.id)
        print(cabang_available)
        bisa_transaksi=cek_expired_kuota(request.user.userprofile.cabang.id)
        print(bisa_transaksi)
        jumlah_grace=cek_jumlah_grace(request.user.userprofile.cabang.id)
        print(jumlah_grace)
        user = User.objects.get(username=request.user.username)
        toko = request.user.userprofile.cabang.nama_toko
        if request.method=="POST":
            try:
                nota = request.GET['nota']
            except:
                nota=""

            kode = request.POST['kode']
            print(kode)
                
            if nota != "":
                penjualan= Penjualan.objects.get(Q(nota=nota) & Q(is_paid=False))
            else:
                cabang = request.user.userprofile.cabang
                penjualan=Penjualan()
                penjualan.cabang=cabang
                penjualan.user=user
                penjualan.save()
                nota=penjualan.nota
            try:
                print(kode)
                if(kode):
                    barang = Barang.objects.get(Q(barcode=kode) & Q(cabang=request.user.userprofile.cabang))
                    try:
                        penjualandetail = PenjualanDetail.objects.get(Q(penjualan=penjualan) & Q(barang=barang))
                        if barang.stok <= 0:
                            messages.add_message(request,messages.SUCCESS,f"Stok {barang.nama} sudah habis.")
                        else:
                            penjualandetail.jumlah += 1
                            penjualandetail.save()
                            barang.stok -= 1
                            barang.save()
                            messages.add_message(request,messages.SUCCESS,f"{barang.nama} berhasil ditambahkan.")
                    except Exception as ex:
                        print(ex)
                        if barang.stok <= 0:
                            messages.add_message(request,messages.SUCCESS,f"Stok {barang.nama} sudah habis.")
                        else:
                            penjualandetail = PenjualanDetail()
                            penjualandetail.jumlah=1
                            penjualandetail.penjualan=penjualan
                            penjualandetail.barang=barang
                            penjualandetail.save()
                            barang.stok -= 1
                            barang.save()
                            messages.add_message(request,messages.SUCCESS,f"{barang.nama} berhasil ditambahkan.")

                    return HttpResponseRedirect(f'/?nota={nota}')
                else:
                    messages.add_message(request,messages.SUCCESS,"Maaf barang dengan kode tersebut belum ada.")    
            except Exception as ex:
                if (kode):
                    messages.add_message(request,messages.SUCCESS,"Maaf barang dengan kode tersebut belum ada.")
                    penjualandetail = PenjualanDetail.objects.all().filter(penjualan=penjualan)
                    if len(penjualandetail)==0:
                        penjualan.delete()
                        return HttpResponseRedirect('/')
                try:
                    return HttpResponseRedirect(f'/?nota={nota}')
                except:
                    messages.add_message(request,messages.SUCCESS,"Silakan login untuk bisa melakukan transaksi.")
                    return HttpResponseRedirect('/')
    else:
        user=None
        is_paket_gratis=False
        is_paket_korporasi=False
        pool_kuota_korporasi=0
        owner_session=False
        korporasi_unlicensed=False

    promo_list = Promo.objects.all().filter(Q(end_period__gte=datetime.datetime.now()) & Q(is_active=True) & Q(kuota__gt=0))

    try:
        if request.method != "POST":
            print('ini bukan post')
            # diubah supaya bs fleksible
            try:
                nota = request.GET['nota']
            except:
                nota=""
        else:
            print('ini post')
            print(f'nota: {nota}')
        
        if(nota!=""):
            penjualan = Penjualan.objects.get(Q(nota=nota) & Q(is_paid=False))
            penjualandetail = PenjualanDetail.objects.all().filter(penjualan=penjualan)
            total = int(penjualan.total)
            jumlah_item = penjualan.items
        else:
            penjualandetail=None
            total=0
            jumlah_item=0
    except Exception as ex:
        return HttpResponseRedirect('/')
    try:
        barang = Barang.objects.all().filter(cabang=request.user.userprofile.cabang, stok__gt=0)
        jml_barang = barang.count()
    except:
        barang = None
        jml_barang=0

    penjualan_pending = Penjualan.objects.all().filter(Q(is_paid=False) & Q(user=user)).order_by('-created_at')
    jumlah_penjualan_pending = penjualan_pending.count()

    bisnis_kecil = DaftarPaket.objects.filter(nama="Bisnis Kecil").first()
    bisnis_medium = DaftarPaket.objects.filter(nama="Bisnis Medium").first()

    # Status add-on untuk popup konfirmasi di landing page
    addon_status = {}
    try:
        from payment.models import TokoAddon
        if request.user.is_authenticated and hasattr(request.user, 'userprofile'):
            _cabang = request.user.userprofile.cabang
            for _type in [TokoAddon.ADDON_BARCODE, TokoAddon.ADDON_NOTA, TokoAddon.ADDON_AKUNTING]:
                _a = TokoAddon.get_for_cabang(_cabang, _type)
                if _a and _a.is_active:
                    addon_status[_type] = {
                        'aktif': True,
                        'expired_at': _a.expired_at.strftime('%d/%m/%Y') if _a.expired_at else '-',
                        'nama_toko': _cabang.nama_toko,
                    }
    except Exception:
        pass

    testimoni = Testimoni.objects.all().filter(is_verified=True)

    print(f"nota: {nota}")

    try:
        is_checked = Penjualan.objects.get(Q(nota=nota) & Q(is_paid=False)).is_checked if nota else False
    except Exception:
        is_checked = False

    context = {
        'penjualandetail':penjualandetail,
        'nota':nota,
        'total':total,
        'jumlah_item':jumlah_item,
        'barang':barang,
        'jml_barang':jml_barang,
        'toko':toko,
        'tanggal':tanggal,
        'penjualan_pending':penjualan_pending,
        'jumlah_penjualan_pending':jumlah_penjualan_pending,
        'bisnis_kecil':bisnis_kecil,
        'bisnis_medium':bisnis_medium,
        'cabang_available':cabang_available,
        'jumlah_grace':jumlah_grace,
        'bisa_transaksi':bisa_transaksi,
        'promo_list':promo_list,
        'jumlah_transaksi':jumlah_transaksi,
        'image_tutorial':image,
        'list_testimoni':testimoni,
        'limapenjualan':limapenjualan,
        'is_paket_gratis':is_paket_gratis,
        'is_paket_korporasi':is_paket_korporasi,
        'pool_kuota_korporasi':pool_kuota_korporasi,
        'owner_session':owner_session,
        'is_checked':is_checked,
        'addon_status': addon_status,
    }
    return render(request,'pos/pos.html',context)

def hapusItems(request):
    try:
        page=request.META['HTTP_REFERER']
    except:
        page="/"
    
    if request.user.is_authenticated and request.user.userprofile.is_active:
        try:
            nota = request.GET['nota']
            id_barang = request.GET['id']
            try:
                penjualan = Penjualan.objects.get(Q(nota=nota) & Q(is_paid=False))
                barang = Barang.objects.get(Q(id=int(id_barang)) & Q(cabang=request.user.userprofile.cabang))
                penjualandetail = PenjualanDetail.objects.get(Q(penjualan=penjualan) & Q(barang=barang))
                # Kembalikan stok sesuai jumlah yang ada di transaksi
                barang.stok += penjualandetail.jumlah
                barang.save()
                penjualandetail.delete()
                _reset_is_checked(nota)

                penjualandetail = PenjualanDetail.objects.all().filter(Q(penjualan=penjualan))
                if len(penjualandetail)==0:
                    penjualan.delete()
                    messages.add_message(request,messages.SUCCESS,f'Silakan bertransaksi kembali dengan Tambah Barang.')
                    return HttpResponseRedirect('/')
                else:
                    messages.add_message(request,messages.SUCCESS,f'Jumlah {barang.nama} telah dihapus.')
                
            except Exception as ex:
                print(ex)
                messages.add_message(request,messages.SUCCESS,'Kode Barang atau Nomor Nota tidak sesuai.')
            return HttpResponseRedirect(page)    
        except:
            messages.add_message(request,messages.SUCCESS,'Data kode barang atau nota belum dimasukkan.')
            nota = None
            id_barang = None
            return HttpResponseRedirect(page)
    else:
        messages.add_message(request,messages.SUCCESS,'Pengguna telah dinonaktifkan. Silakan menghubungi pemilik toko untuk konfirmasi.')
        return HttpResponseRedirect(page)

def tambahItems(request):
    try:
        page=request.META['HTTP_REFERER']
    except:
        page="/"
    
    if request.user.is_authenticated and request.user.userprofile.is_active:
        try:
            nota = request.GET['nota']
            id_barang = request.GET['id']
            try:
                if nota == "":
                    penjualan = Penjualan()
                else:
                    penjualan = Penjualan.objects.get(nota=nota)
                nota = penjualan.nota
                barang = Barang.objects.get(Q(id=int(id_barang)) & Q(cabang=request.user.userprofile.cabang))
                penjualandetail = PenjualanDetail.objects.get(Q(penjualan=penjualan) & Q(barang=barang) & Q(is_paid=False))
                if barang.stok <= 0:
                    messages.add_message(request,messages.SUCCESS,f'Stok {barang.nama} tersedia hanya {penjualandetail.jumlah} (sudah habis).')
                else:
                    penjualandetail.jumlah += 1
                    penjualandetail.save()
                    barang.stok -= 1
                    barang.save()
                    _reset_is_checked(nota)
                    messages.add_message(request,messages.SUCCESS,f'Jumlah {barang.nama} telah ditambah 1.')
                return HttpResponseRedirect(f"/?nota={nota}")
            except:
                messages.add_message(request,messages.SUCCESS,'Kode Barang atau Nomor Nota tidak sesuai.')
            return HttpResponseRedirect(page)    
        except:
            messages.add_message(request,messages.SUCCESS,'Data kode barang atau nota belum dimasukkan.')
            nota = None
            id_barang = None
    else:
        messages.add_message(request,messages.SUCCESS,'Pengguna telah dinonaktifkan. Silakan menghubungi pemilik toko untuk konfirmasi.')
    return HttpResponseRedirect(page)

def ubahItems(request):
    try:
        page=request.META['HTTP_REFERER']
    except:
        page="/"
    
    if request.user.is_authenticated and request.user.userprofile.is_active:
        try:
            nota = request.GET['nota']
            id_barang = int(request.GET['id'])
            jumlah = int(request.POST['jumlah'])
            try:
                penjualan = Penjualan.objects.get(Q(nota=nota) & Q(is_paid=False))
                barang = Barang.objects.get(Q(id=int(id_barang)) & Q(cabang=request.user.userprofile.cabang))
                penjualandetail = PenjualanDetail.objects.get(Q(penjualan=penjualan) & Q(barang=barang))
                jumlah_lama = penjualandetail.jumlah
                if int(jumlah) == 0:
                    # Hapus item, kembalikan stok
                    barang.stok += jumlah_lama
                    barang.save()
                    penjualandetail.delete()
                    penjualan.items -= 1
                    penjualan.save()
                    messages.add_message(request,messages.SUCCESS,f'{barang.nama} dihapus karena dalam daftar menjadi nol.')
                else:
                    delta = int(jumlah) - jumlah_lama
                    if delta > 0 and barang.stok < delta:
                        # Stok tidak cukup — set ke maksimal yang bisa
                        jumlah_bisa = jumlah_lama + barang.stok
                        messages.add_message(request,messages.SUCCESS,
                            f'Stok {barang.nama} tersedia hanya {jumlah_bisa}, jumlah diset ke {jumlah_bisa}.')
                        penjualandetail.jumlah = jumlah_bisa
                        barang.stok = 0
                    else:
                        penjualandetail.jumlah = int(jumlah)
                        barang.stok -= delta
                        messages.add_message(request,messages.SUCCESS,f'Jumlah {barang.nama} telah diupdate menjadi {jumlah}.')
                    penjualandetail.save()
                    barang.save()
                    _reset_is_checked(nota)

                if penjualan.items == 0:
                    penjualan.delete()
                    return HttpResponseRedirect('/')

            except Exception as ex:
                print(ex)
                messages.add_message(request,messages.SUCCESS,'Kode Barang atau Nomor Nota tidak sesuai.')
                return HttpResponseRedirect('/')
            return HttpResponseRedirect(page)    
        except Exception as ex:
            print(ex)
            messages.add_message(request,messages.SUCCESS,'Data kode barang atau nota belum dimasukkan.')
            nota = None
            id_barang = None
    else:
        messages.add_message(request,messages.SUCCESS,'Pengguna telah dinonaktifkan. Silakan menghubungi pemilik toko untuk konfirmasi.')
    return HttpResponseRedirect(page)

def tambahBarang(request):
    try:
        page=request.META['HTTP_REFERER']
    except:
        page="/"

    if request.user.is_authenticated and request.user.userprofile.is_active:
        user = User.objects.get(username=request.user.username)
        cabang = request.user.userprofile.cabang
        print(cabang.id)
        try:
            nota = request.GET['nota']
        except:
            messages.add_message(request,messages.SUCCESS,'Id Nota belum ada...')
            return HttpResponseRedirect("/")

        try:
            id_barang = request.GET['id']
        except Exception as ex:
            print(ex)
            # messages.add_message(request,messages.SUCCESS,'Kode Barang belum ada.')
            return HttpResponseRedirect(f"/?nota={nota}")
        
        print(nota)
        
        if(nota):
            try:
                penjualan = Penjualan.objects.get(Q(nota=nota) & Q(is_paid=False))
                try:
                    barang = Barang.objects.get(Q(id=int(id_barang)) & Q(cabang = request.user.userprofile.cabang))
                except:
                    messages.add_message(request,messages.SUCCESS,'Kode Barang tidak ditemukan.')
                    return HttpResponseRedirect(f"/?nota={nota}")   
                
                try:
                    penjualandetail = PenjualanDetail.objects.get(Q(penjualan=penjualan) & Q(barang=barang))
                    if barang.stok <= 0:
                        messages.add_message(request,messages.SUCCESS,f'Stok {barang.nama} tersedia hanya {penjualandetail.jumlah} (sudah habis).')
                    else:
                        penjualandetail.jumlah += 1
                        penjualandetail.save()
                        barang.stok -= 1
                        barang.save()
                        _reset_is_checked(penjualan.nota)
                        messages.add_message(request,messages.SUCCESS,f'Jumlah {barang.nama} telah ditambah 1.')
                except:
                    if barang.stok <= 0:
                        messages.add_message(request,messages.SUCCESS,f'Stok {barang.nama} sudah habis.')
                    else:
                        penjualandetail = PenjualanDetail()
                        penjualandetail.barang=barang
                        penjualandetail.jumlah=1
                        penjualandetail.penjualan=penjualan
                        penjualandetail.save()
                        barang.stok -= 1
                        barang.save()
                        _reset_is_checked(penjualan.nota)
                        messages.add_message(request,messages.SUCCESS,f'Jumlah {barang.nama} telah ditambah 1.')
            except:
                messages.add_message(request,messages.SUCCESS,'Kode Barang atau Nomor Nota tidak sesuai.')
                return HttpResponseRedirect('/')
            return HttpResponseRedirect(f"?nota={nota}")
        else:
            penjualan = Penjualan()
            penjualan.cabang=cabang
            penjualan.user=user
            penjualan.save()
            barang = Barang.objects.get(Q(id=int(id_barang)) & Q(cabang=request.user.userprofile.cabang))
            if barang.stok <= 0:
                messages.add_message(request,messages.SUCCESS,f'Stok {barang.nama} sudah habis.')
                penjualan.delete()
            else:
                penjualandetail = PenjualanDetail()
                penjualandetail.penjualan=penjualan
                penjualandetail.jumlah=1
                penjualandetail.barang=barang
                penjualandetail.save()
                barang.stok -= 1
                barang.save()
                messages.add_message(request,messages.SUCCESS,f'Jumlah {barang.nama} telah ditambah 1.')
    else:
        messages.add_message(request,messages.SUCCESS,'Pengguna telah dinonaktifkan. Silakan menghubungi pemilik toko untuk konfirmasi.')
    try:
        return HttpResponseRedirect(f"/?nota={penjualan.nota}")
    except:
        return HttpResponseRedirect('/')
    
def hapusTransaksi(request):
    try:
        page=request.META['HTTP_REFERER']
    except:
        page="/"
    if request.user.is_authenticated and request.user.userprofile.is_active:
        print('hallo')
        try:
            nota = request.GET['nota']
            try:
                penjualan_hapus = Penjualan.objects.get(Q(nota=nota) & Q(user=request.user) & Q(is_paid=False))
                for detail in PenjualanDetail.objects.filter(penjualan=penjualan_hapus):
                    detail.barang.stok += detail.jumlah
                    detail.barang.save()
                penjualan_hapus.delete()
                messages.success(request, 'Transaksi berhasil dihapus.')
                return HttpResponseRedirect(page)
            except Exception as ex:
                print(ex)
                messages.add_message(request,messages.SUCCESS,'Nota Penjualan tidak diketemukan...')
                return HttpResponseRedirect(page)
        except Exception as ex:
            print(ex)
            messages.add_message(request,messages.SUCCESS,'Nota Penjualan tidak diketemukan...')
            return HttpResponseRedirect(page)
    else:
        messages.add_message(request,messages.SUCCESS,'Pengguna telah dinonaktifkan. Silakan menghubungi pemilik toko untuk konfirmasi.')
        return HttpResponseRedirect(page)
        
def loginkan(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return HttpResponseRedirect('/management/')
        # Owner Korporasi (tidak punya userprofile) → dashboard owner
        if hasattr(request.user, 'owner_profile'):
            return HttpResponseRedirect('/owner/')
        return HttpResponseRedirect('/')
    else:
        if request.method=="POST":
            username = str(request.POST['username']).lower()
            password = request.POST['password']
            user = authenticate(username=username,password=password)
            if(user):
                if user.is_staff:
                    login(request, user)
                    _hapus_session_lain(user, request.session.session_key)
                    messages.add_message(request, messages.SUCCESS, f"Selamat datang {user.username}")
                    return HttpResponseRedirect('/management/')
                if user.userprofile.is_active:
                    pending_dihapus = _hapus_transaksi_pending_user(user)
                    login(request,user)
                    _hapus_session_lain(user, request.session.session_key)
                    messages.add_message(request,messages.SUCCESS,f"Selamat datang {user.userprofile.nama_lengkap}")
                    addLog(user,user.userprofile.cabang,"login",f"login berhasil")
                    if pending_dihapus:
                        addLog(user, user.userprofile.cabang, "login", f"menghapus {pending_dihapus} transaksi pending milik kasir saat login")
                    return HttpResponseRedirect('/')
                else:
                    messages.add_message(request,messages.SUCCESS,f"Pengguna {user.userprofile.nama_lengkap} telah dinonaktifkan. Silakan menghubungi Pemilik toko untuk konfirmasi.")
                    addLog(user,user.userprofile.cabang,"login",f"pengguna dinonaktifkan, tidak bisa login.")
                    return HttpResponseRedirect('/')
            else:
                # cek di cabang ada atau tidak.
                try:
                    cabang = Cabang.objects.get(email=username)
                    username=f"{cabang.prefix}1"
                    user=authenticate(username=username,password=password)
                    if(user):
                        pending_dihapus = _hapus_transaksi_pending_user(user)
                        login(request,user)
                        _hapus_session_lain(user, request.session.session_key)
                        messages.add_message(request,messages.SUCCESS,f"Selamat datang {user.userprofile.nama_lengkap}")
                        addLog(user,user.userprofile.cabang,"login",f"login berhasil")
                        if pending_dihapus:
                            addLog(user, user.userprofile.cabang, "login", f"menghapus {pending_dihapus} transaksi pending milik kasir saat login")
                        return HttpResponseRedirect('/')
                except:
                    pass            
                addLog("","","login",f"login dengan user {username} dan password {password} gagal")
                messages.add_message(request,messages.SUCCESS,f"Username dan Password tidak sesuai, silakan ulangi kembali.")
        try:
            toko = request.user.userprofile.cabang.nama_toko
        except:
            toko=""
        context = {
            'toko':toko
        }
        return render(request,'pos/login.html',context)
    
def logoutkan(request):
    try:
        pending_dihapus = _hapus_transaksi_pending_user(request.user)
        if pending_dihapus:
            addLog(
                request.user,
                request.user.userprofile.cabang,
                "logout",
                f"logout berhasil dan menghapus {pending_dihapus} transaksi pending",
            )
        else:
            addLog(request.user,request.user.userprofile.cabang,"logout",f"logout berhasil")
    except:
        try:
            addLog(request.user,request.user.userprofile.cabang,"logout",f"logout berhasil")
        except:
            pass
    logout(request)
    return HttpResponseRedirect('/')

def bayarTransaksi(request):
    if request.user.is_authenticated and request.user.userprofile.is_active:
        try:
            nota = request.GET['nota']
            try:
                penjualan = Penjualan.objects.get(Q(nota=nota) & Q(is_paid=False))
                if request.method=="POST":
                    try:
                        metode_dipilih = int(request.POST['metode'])
                        cabang_obj = request.user.userprofile.cabang
                        is_gratis = cabang_obj.paket is None and cabang_obj.owner_id is None

                        # Tolak tempo untuk paket gratis
                        if metode_dipilih == 2 and is_gratis:
                            messages.add_message(request, messages.SUCCESS, 'Pembayaran tempo tidak tersedia untuk paket gratis.')
                            return HttpResponseRedirect(f'/?nota={nota}')

                        no_nota = penjualan.cabang.no_nota
                        penjualan.is_paid = True
                        penjualan.metode = metode_dipilih
                        penjualan.customer = request.POST['pembeli']
                        penjualan.no_nota = str(no_nota).zfill(5)

                        if metode_dipilih == 2:
                            # Tempo: tgl_bayar dikosongkan, simpan jatuh_tempo
                            import datetime as _dt
                            jatuh_tempo_str = request.POST.get('jatuh_tempo', '')
                            try:
                                penjualan.jatuh_tempo = _dt.date.fromisoformat(jatuh_tempo_str)
                            except ValueError:
                                messages.add_message(request, messages.SUCCESS, 'Tanggal jatuh tempo tidak valid.')
                                return HttpResponseRedirect(f'/?nota={nota}')
                            penjualan.tgl_bayar = None
                            penjualan.is_tempo_lunas = False
                        else:
                            penjualan.tgl_bayar = datetime.datetime.now()

                        penjualan.save()

                        cabangnya = penjualan.cabang
                        cabangnya.no_nota += 1
                        cabangnya.save()

                        penjualandetail = PenjualanDetail.objects.filter(
                            penjualan=penjualan
                        ).select_related('barang')
                        for detail in penjualandetail:
                            detail.barang.jumlah_dibeli += detail.jumlah
                            detail.barang.save(update_fields=['jumlah_dibeli'])

                        addLog(request.user, request.user.userprofile.cabang, "pembayaran",
                               f"{'Pembayaran Tempo' if metode_dipilih == 2 else 'Pembayaran'} no. transaksi: {penjualan.nota}")

                        cabang = Cabang.objects.get(id=request.user.userprofile.cabang.id)
                        if cabang.owner_id:
                            if _korporasi_aktif(cabang):
                                owner = cabang.owner
                                if owner.kuota_transaksi_pool > 0:
                                    owner.kuota_transaksi_pool -= 1
                                    owner.save()
                            else:
                                if cabang.kuota_transaksi > 0:
                                    cabang.kuota_transaksi -= 1
                                    cabang.save()
                        else:
                            cabang.kuota_transaksi -= 1
                            cabang.save()

                        import time as _time
                        _cabang_qos = penjualan.cabang
                        if _cabang_qos.paket is None and _cabang_qos.owner_id is None:
                            _time.sleep(1.5)
                        elif _cabang_qos.paket is not None:
                            _time.sleep(0.5)

                        if metode_dipilih == 2:
                            return HttpResponseRedirect(f'/print/tempo/{nota}')
                        return HttpResponseRedirect(f'/print/{nota}')
                    except Exception as ex:
                        print(ex)
                        messages.add_message(request,messages.SUCCESS,'Pembayaran Gagal Dilakukan.')
            except Exception as ex:
                # print(ex)
                messages.add_message(request,messages.SUCCESS,"Nomor nota tidak diketemukan.")    
                nota=None
        except Exception as ex:
            # print(ex)
            nota=None
            messages.add_message(request,messages.SUCCESS,"Nomor nota kosong.")
        
        if nota:
            if request.method=="POST":
                print(request.POST)
    else:
        messages.add_message(request,messages.SUCCESS,"Pengguna telah dinonaktifkan dan tidak bisa melakukan transaksi. silakan hubungi Pemilik Toko.")
    return HttpResponseRedirect('/')

def printTransaksi(request,nota):
    if request.user.is_authenticated and request.user.userprofile.is_active:
        try:
            penjualan = Penjualan.objects.get(Q(nota=nota) & Q(user=request.user)  & Q(is_void=False))
            penjualandetail = PenjualanDetail.objects.all().filter(penjualan=penjualan)
            cabang = request.user.userprofile.cabang
            nama_toko = cabang.nama_toko
            alamat_toko = cabang.alamat_toko
            telpon_toko = cabang.telpon
            total = int(penjualan.total)

            from payment.models import TokoAddon, AddonConfig
            nota_addon = TokoAddon.get_for_cabang(cabang, TokoAddon.ADDON_NOTA)
            nota_config = {}
            if nota_addon:
                cfg_obj = AddonConfig.objects.filter(cabang=cabang, addon_type=TokoAddon.ADDON_NOTA).first()
                if cfg_obj:
                    nota_config = cfg_obj.config

            nama_lengkap = penjualan.user.userprofile.nama_lengkap.strip()
            kasir_nama = nama_lengkap.split()[0] if nama_lengkap else penjualan.user.username

            if penjualan.metode==0:
                bayar="cash"
            else:
                bayar="transfer"
            context = {
                'penjualan':penjualan,
                'penjualandetail':penjualandetail,
                'nama_toko':nama_toko,
                'alamat_toko':alamat_toko,
                'telpon_toko':telpon_toko,
                'bayar':bayar,
                'total':total,
                'kasir_nama': kasir_nama,
                'nota_config': nota_config,
                'nota_addon_aktif': bool(nota_addon),
            }
            messages.add_message(request,messages.SUCCESS,"Pembayaran Berhasil.")
            return render(request,'pos/print.html',context)
        except Exception as ex:
            messages.add_message(request,messages.SUCCESS,'Nota yang akan dicetak tidak diketemukan..')
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login untuk bisa mencetak transaksi.")
        return HttpResponseRedirect('/')
    
def printCekTransaksi(request, nota):
    if request.user.is_authenticated and request.user.userprofile.is_active:
        cabang = request.user.userprofile.cabang
        if cabang.paket is None and cabang.owner_id is None:
            messages.add_message(request, messages.SUCCESS, 'Fitur Cek Barang hanya tersedia untuk paket berbayar.')
            return HttpResponseRedirect(f'/?nota={nota}')
        try:
            penjualan = Penjualan.objects.get(Q(nota=nota) & Q(user=request.user) & Q(is_paid=False))
            penjualan.is_checked = True
            penjualan.save(update_fields=['is_checked'])
            penjualandetail = PenjualanDetail.objects.filter(penjualan=penjualan)
            nama_toko = request.user.userprofile.cabang.nama_toko
            alamat_toko = request.user.userprofile.cabang.alamat_toko
            telpon_toko = request.user.userprofile.cabang.telpon
            total = int(penjualan.total)
            context = {
                'penjualan': penjualan,
                'penjualandetail': penjualandetail,
                'nama_toko': nama_toko,
                'alamat_toko': alamat_toko,
                'telpon_toko': telpon_toko,
                'total': total,
            }
            return render(request, 'pos/print_cek.html', context)
        except Exception as ex:
            messages.add_message(request, messages.SUCCESS, 'Nota yang akan dicetak tidak diketemukan.')
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request, messages.SUCCESS, "Silakan Login untuk bisa mencetak.")
        return HttpResponseRedirect('/')


def printTempoTransaksi(request, nota):
    if request.user.is_authenticated and request.user.userprofile.is_active:
        try:
            penjualan = Penjualan.objects.get(
                Q(nota=nota) & Q(user=request.user) & Q(metode=2) & Q(is_paid=True)
            )
            penjualandetail = PenjualanDetail.objects.filter(penjualan=penjualan)
            nama_toko = request.user.userprofile.cabang.nama_toko
            alamat_toko = request.user.userprofile.cabang.alamat_toko
            telpon_toko = request.user.userprofile.cabang.telpon
            total = int(penjualan.total)
            context = {
                'penjualan': penjualan,
                'penjualandetail': penjualandetail,
                'nama_toko': nama_toko,
                'alamat_toko': alamat_toko,
                'telpon_toko': telpon_toko,
                'total': total,
            }
            messages.add_message(request, messages.SUCCESS, 'Transaksi tempo berhasil dicatat.')
            return render(request, 'pos/print_tempo.html', context)
        except Exception:
            messages.add_message(request, messages.SUCCESS, 'Nota tempo tidak ditemukan.')
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request, messages.SUCCESS, 'Silakan Login untuk mencetak.')
        return HttpResponseRedirect('/')


def gantiStatusOpen(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            try:
                id = request.GET['id']
                nota = request.GET['nota']
                status = int(request.GET['status'])

                penjualandetail = PenjualanDetail.objects.get(id=id)
                if status == 0:
                    penjualandetail.is_open=False
                    messages.add_message(request,messages.SUCCESS,f"Harga barang {penjualandetail.barang.nama} mengunakan metode harga normal kembali.")
                elif status ==1:
                    penjualandetail.is_open=True
                    
                    messages.add_message(request,messages.SUCCESS,f"Harga barang {penjualandetail.barang.nama} mengunakan metode edit harga, dengan minimal harga Rp.1,00.")
                    messages.add_message(request,messages.SUCCESS,"Harga akan update otomatis diupdate oleh sistem 5 Detik setelah Sobat melakukan pengetikan harga. ")
                    messages.add_message(request,messages.SUCCESS,"Untuk mengembalikan ke harga semula silakan tekan tombol refresh berwarna merah di samping. ")
                penjualandetail.save()

                return HttpResponseRedirect(f'/?nota={nota}')
            except:
                return HttpResponseRedirect('/')
        else:
            messages.add_message(request,messages.SUCCESS,"Hanya Admin yang berhak melakukan update harga.")
            return HttpResponseRedirect('/')    
        
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login untuk bisa mencetak transaksi.")
        return HttpResponseRedirect('/')
    

def updateBarangSatuan(request):
    try:
        page=request.META['HTTP_REFERER']
    except:
        page="/"
    
    if request.user.is_authenticated and request.user.userprofile.is_active:
        try:
            nota = request.GET['nota']
            id_barang = int(request.GET['id'])
            update_harga = int(request.GET['harga_baru'])
            if update_harga>0:
                try:
                    penjualan = Penjualan.objects.get(Q(nota=nota) & Q(is_paid=False))
                    penjualandetail = PenjualanDetail.objects.get(Q(penjualan=penjualan) & Q(id=id_barang))
                    harga_awal = penjualandetail.harga
                    penjualandetail.harga = update_harga
                    penjualandetail.save()
                    messages.add_message(request,messages.SUCCESS,f"Harga barang {penjualandetail.barang.nama} sudah diupdate dari harga asli: Rp.{int(harga_awal):,} menjadi harga Rp.{int(update_harga):,}.")
                    messages.add_message(request,messages.SUCCESS,"Untuk mengembalikan ke harga semula silakan tekan tombol refresh berwarna merah di samping.")
                    messages.add_message(request,messages.SUCCESS,"Silakan Sobat Cek terlebih dahulu 1 per 1 harga manual sebelum melakukan pembayaran. Terima kasih.")
                except Exception as ex:
                    print(ex)
                    messages.add_message(request,messages.SUCCESS,'Kode Barang atau Nomor Nota tidak sesuai.')
            else:
                messages.add_message(request,messages.SUCCESS,f'Harga barang tidak boleh nol. silakan cek input Sobat.')
            return HttpResponseRedirect(page)    
        except Exception as ex:
            print(ex)
            messages.add_message(request,messages.SUCCESS,'Data kode barang atau nota belum dimasukkan.')
            nota = None
            id_barang = None
    else:
        messages.add_message(request,messages.SUCCESS,'Silakan login untuk bisa melakukan transaksi...')
    return HttpResponseRedirect(page)

def reprintTransaksi(request,nota):
    if request.user.is_authenticated  and request.user.userprofile.is_active:
        try:
            penjualan = Penjualan.objects.get(Q(nota=nota) & Q(user=request.user)  & Q(is_void=False))
            penjualandetail = PenjualanDetail.objects.all().filter(penjualan=penjualan)
            cabang = request.user.userprofile.cabang
            nama_toko = cabang.nama_toko
            alamat_toko = cabang.alamat_toko
            telpon_toko = cabang.telpon
            total = int(penjualan.total)
            penjualan.reprint_nota += 1
            penjualan.save()

            from payment.models import TokoAddon, AddonConfig
            nota_addon = TokoAddon.get_for_cabang(cabang, TokoAddon.ADDON_NOTA)
            nota_config = {}
            if nota_addon:
                cfg_obj = AddonConfig.objects.filter(cabang=cabang, addon_type=TokoAddon.ADDON_NOTA).first()
                if cfg_obj:
                    nota_config = cfg_obj.config

            nama_lengkap = penjualan.user.userprofile.nama_lengkap.strip()
            kasir_nama = nama_lengkap.split()[0] if nama_lengkap else penjualan.user.username

            if penjualan.metode==0:
                bayar="cash"
            else:
                bayar="transfer"
            context = {
                'penjualan':penjualan,
                'penjualandetail':penjualandetail,
                'nama_toko':nama_toko,
                'alamat_toko':alamat_toko,
                'telpon_toko':telpon_toko,
                'bayar':bayar,
                'total':total,
                'kasir_nama': kasir_nama,
                'nota_config': nota_config,
                'nota_addon_aktif': bool(nota_addon),
            }
            return render(request,'pos/reprint.html',context)
        except Exception as ex:
            messages.add_message(request,messages.SUCCESS,'Nota yang akan dicetak tidak diketemukan..')
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login untuk bisa mencetak transaksi.")
        return HttpResponseRedirect('/')
    

def printKuitansi(request,nota):
    if request.user.is_authenticated and request.user.userprofile.is_active:
        try:
            penjualan = Penjualan.objects.get(Q(nota=nota) & Q(user=request.user) & Q(is_void=False))
            penjualandetail = PenjualanDetail.objects.all().filter(penjualan=penjualan)
            nama_toko = request.user.userprofile.cabang.nama_toko
            alamat_toko = request.user.userprofile.cabang.alamat_toko
            telpon_toko = request.user.userprofile.cabang.telpon
            total = int(penjualan.total)
            penjualan.cetak_kuitansi += 1
            penjualan.save()

            id_pemilik = f"{penjualan.cabang.prefix}1"
            nama_pemilik = User.objects.get(username=id_pemilik).userprofile.nama_lengkap

            if penjualan.metode==0:
                bayar="cash"
            else:
                bayar="transfer"
            context = {
                'penjualan':penjualan,
                'penjualandetail':penjualandetail,
                'nama_toko':nama_toko,
                'alamat_toko':alamat_toko,
                'telpon_toko':telpon_toko,
                'bayar':bayar,
                'total':total,
                'nama_pemilik':nama_pemilik
            }
            return render(request,'pos/print_kuitansi.html',context)
        except Exception as ex:
            messages.add_message(request,messages.SUCCESS,'Nota yang akan dicetak tidak diketemukan..')
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login untuk bisa mencetak transaksi.")
        return HttpResponseRedirect('/')

# ─── Lupa Password ────────────────────────────────────────────────────────────

def lupaPassword(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        try:
            user = User.objects.get(email__iexact=email)
        except User.DoesNotExist:
            # Tetap tampilkan pesan sukses agar email valid tidak terbocorkan
            messages.success(request, "Jika email terdaftar, tautan reset telah dikirim.")
            return HttpResponseRedirect('/lupa-password/')

        # Nonaktifkan token lama yang belum dipakai
        ResetPasswordToken.objects.filter(user=user, is_used=False).update(is_used=True)

        # Buat token baru
        import datetime as _dt
        token = ResetPasswordToken(
            user=user,
            expired_at=_dt.datetime.now() + _dt.timedelta(hours=1)
        )
        token.save()

        reset_url = f"{request.scheme}://{request.get_host()}/reset-password/{token.token}/"
        body = (
            f"Halo {user.username},\n\n"
            f"Kami menerima permintaan reset password untuk akun POSMI Anda.\n\n"
            f"Klik tautan berikut untuk mereset password (berlaku 1 jam):\n"
            f"{reset_url}\n\n"
            f"Jika Anda tidak meminta reset password, abaikan email ini.\n\n"
            f"— Tim POSMI"
        )
        posmiMail("Reset Password POSMI", body, user.email)
        messages.success(request, "Tautan reset password telah dikirim ke email Anda.")
        return HttpResponseRedirect('/lupa-password/')

    return render(request, 'pos/lupa_password.html')


def resetPassword(request, token):
    try:
        reset = ResetPasswordToken.objects.get(token=token)
    except ResetPasswordToken.DoesNotExist:
        messages.error(request, "Tautan tidak valid.")
        return HttpResponseRedirect('/login/')

    if not reset.is_valid:
        messages.error(request, "Tautan sudah tidak berlaku atau telah digunakan.")
        return HttpResponseRedirect('/login/')

    if request.method == 'POST':
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        if len(password1) < 6:
            messages.error(request, "Password minimal 6 karakter.")
            return render(request, 'pos/reset_password.html', {'reset': reset})
        if password1 != password2:
            messages.error(request, "Konfirmasi password tidak cocok.")
            return render(request, 'pos/reset_password.html', {'reset': reset})

        reset.user.set_password(password1)
        reset.user.save()
        reset.is_used = True
        reset.save()
        messages.success(request, f"Password berhasil diubah. Silakan masuk kembali.")
        return HttpResponseRedirect('/login/')

    return render(request, 'pos/reset_password.html', {'reset': reset})


# ─── Riwayat Penjualan Saya ────────────────────────────────────────────────────

def cetakRekap(request):
    """
    Cetak rekap penjualan sebagai PDF/Excel.
    GET params:
      - mode: 'bulanan' | 'harian'
      - bulan: int (1-12)
      - tahun: int
      - tanggal: YYYY-MM-DD (untuk mode harian)
    """
    if not request.user.is_authenticated:
        return HttpResponseRedirect('/login/')

    from django.db.models import Sum
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    now   = datetime.datetime.now()
    mode  = request.GET.get('mode', 'bulanan')
    bulan = int(request.GET.get('bulan', now.month))
    tahun = int(request.GET.get('tahun', now.year))

    BULAN_ID = ['','Januari','Februari','Maret','April','Mei','Juni',
                'Juli','Agustus','September','Oktober','November','Desember']

    cabang = request.user.userprofile.cabang

    if mode == 'harian':
        tanggal_str = request.GET.get('tanggal', '')
        try:
            tgl = datetime.datetime.strptime(tanggal_str, '%Y-%m-%d').date()
        except ValueError:
            messages.add_message(request, messages.SUCCESS, "Format tanggal tidak valid.")
            return HttpResponseRedirect(f'/riwayat/?bulan={bulan}&tahun={tahun}')

        qs = Penjualan.objects.filter(
            cabang=cabang, is_paid=True, is_void=False,
            tgl_bayar__date=tgl
        ).prefetch_related('penjualandetail_set__barang').order_by('tgl_bayar')
        judul = f"Rekap Harian {tgl.strftime('%d %B %Y')}"
        nama_file = f"rekap_harian_{tgl.strftime('%Y%m%d')}_{cabang.prefix}.xlsx"
    else:
        qs = Penjualan.objects.filter(
            cabang=cabang, is_paid=True, is_void=False,
            tgl_bayar__month=bulan, tgl_bayar__year=tahun
        ).prefetch_related('penjualandetail_set__barang').order_by('tgl_bayar')
        judul = f"Rekap Bulanan {BULAN_ID[bulan]} {tahun}"
        nama_file = f"rekap_{tahun}{bulan:02d}_{cabang.prefix}.xlsx"

    # ── Build Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap"

    hd_fill   = PatternFill(start_color="1E2C5A", end_color="1E2C5A", fill_type="solid")
    hd_font   = Font(bold=True, color="FFFFFF", size=11)
    sub_fill  = PatternFill(start_color="F0F4FB", end_color="F0F4FB", fill_type="solid")
    alt_fill  = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin = Border(
        left=Side(style='thin', color='D8E3F5'),
        right=Side(style='thin', color='D8E3F5'),
        top=Side(style='thin', color='D8E3F5'),
        bottom=Side(style='thin', color='D8E3F5'),
    )
    center = Alignment(horizontal='center', vertical='center')

    # Judul
    ws.merge_cells('A1:G1')
    ws['A1'] = judul
    ws['A1'].font = Font(bold=True, size=13, color="1E2C5A")
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Toko: {cabang.nama_toko} ({cabang.nama_cabang}) · Dicetak: {now.strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(size=10, color="8899BB")
    ws.row_dimensions[1].height = 22
    ws.append([])

    # Header kolom
    headers = ['No. Nota', 'Tanggal', 'Waktu', 'Customer', 'Metode', 'Total', 'Items']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = hd_font
        cell.fill = hd_fill
        cell.alignment = center
        cell.border = thin

    col_widths = [14, 14, 8, 20, 10, 16, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.row_dimensions[4].height = 18
    ws.freeze_panes = 'A5'

    total_grand = 0
    row = 5
    for idx, t in enumerate(qs):
        fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
        data = [
            t.no_nota,
            t.tgl_bayar.strftime('%d/%m/%Y') if t.tgl_bayar else '-',
            t.tgl_bayar.strftime('%H:%M') if t.tgl_bayar else '',
            t.customer or '-',
            'Cash' if t.metode == 0 else 'Transfer',
            int(t.total),
            t.items,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if fill.fill_type:
                cell.fill = fill
            if col in (6,):
                cell.alignment = Alignment(horizontal='right')
            elif col in (1,2,3,5,7):
                cell.alignment = center
        total_grand += int(t.total)
        row += 1

    # Total row
    ws.cell(row=row, column=5, value="TOTAL").font = Font(bold=True, color="1E2C5A")
    total_cell = ws.cell(row=row, column=6, value=total_grand)
    total_cell.font = Font(bold=True, color="1E2C5A")
    total_cell.alignment = Alignment(horizontal='right')
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = sub_fill
        ws.cell(row=row, column=col).border = thin

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nama_file}"'
    wb.save(response)
    return response


def riwayatSaya(request):
    if not request.user.is_authenticated:
        messages.add_message(request, messages.SUCCESS, "Silakan login terlebih dahulu.")
        return HttpResponseRedirect('/login/')

    from django.core.paginator import Paginator
    from django.db.models import Sum, Count

    now = datetime.datetime.now()

    cabang = request.user.userprofile.cabang
    is_owner_view = request.user.is_superuser

    # Ambil bulan & tahun yang punya transaksi (paid)
    periode_base = Penjualan.objects.filter(is_paid=True, tgl_bayar__isnull=False)
    if is_owner_view:
        periode_base = periode_base.filter(cabang=cabang)
    else:
        periode_base = periode_base.filter(user=request.user)
    periode_list = periode_base.dates('tgl_bayar', 'month', order='DESC')

    # Parameter filter
    bulan_param = int(request.GET.get('bulan', now.month))
    tahun_param = int(request.GET.get('tahun', now.year))
    per_page    = int(request.GET.get('per_page', 25))
    if per_page not in [25, 50, 75, 100]:
        per_page = 25

    # Query transaksi selesai sesuai filter
    paid_base = Penjualan.objects.filter(
        is_paid=True,
        is_void=False,
        tgl_bayar__month=bulan_param,
        tgl_bayar__year=tahun_param,
    )
    if is_owner_view:
        qs_paid = paid_base.filter(cabang=cabang).select_related('user__userprofile').order_by('-tgl_bayar')
    else:
        qs_paid = paid_base.filter(user=request.user).order_by('-tgl_bayar')

    # Transaksi pending (belum dibayar)
    pending_base = Penjualan.objects.filter(is_paid=False)
    if is_owner_view:
        qs_pending = pending_base.filter(cabang=cabang).select_related('user__userprofile').order_by('-created_at')
    else:
        qs_pending = pending_base.filter(user=request.user).order_by('-created_at')

    # Summary
    total_penjualan = qs_paid.aggregate(total=Sum('total'))['total'] or 0
    jumlah_pembeli  = qs_paid.exclude(customer='').exclude(customer__isnull=True).count()
    jumlah_transaksi = qs_paid.count()

    # Pagination
    paginator = Paginator(qs_paid, per_page)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    # Tanggal yang punya transaksi di bulan yang dipilih (untuk pilihan rekap harian)
    tanggal_bertransaksi = list(
        Penjualan.objects
        .filter(
            cabang=cabang,
            is_paid=True, is_void=False,
            tgl_bayar__month=bulan_param,
            tgl_bayar__year=tahun_param,
        )
        .exclude(tgl_bayar=None)
        .dates('tgl_bayar', 'day', order='ASC')
    )

    # Nama bulan helper
    BULAN = ['','Januari','Februari','Maret','April','Mei','Juni',
             'Juli','Agustus','September','Oktober','November','Desember']

    return render(request, 'pos/riwayat_saya.html', {
        'page_obj':             page_obj,
        'periode_list':         periode_list,
        'bulan_param':          bulan_param,
        'tahun_param':          tahun_param,
        'per_page':             per_page,
        'total_penjualan':      total_penjualan,
        'jumlah_pembeli':       jumlah_pembeli,
        'jumlah_transaksi':     jumlah_transaksi,
        'qs_pending':           qs_pending,
        'nama_bulan':           BULAN[bulan_param],
        'BULAN':                BULAN,
        'tanggal_bertransaksi': tanggal_bertransaksi,
        'is_owner_view':        is_owner_view,
    })

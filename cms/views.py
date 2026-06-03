from django.shortcuts import render, HttpResponseRedirect, HttpResponse, get_object_or_404
from django.contrib import messages
from pos.models import Penjualan, Barang, DetailWalet
from django.db.models import Q,Sum,Count
import datetime
from django.contrib.auth.models import User
from django.core.files.storage import default_storage
from promo.models import Promo, PromoUsed
from .forms import FormInfoToko, FormUserProfile, FormUser, FormBarang
from stock.models import Cabang,UserProfile,DaftarPaket
from django.urls import reverse
from django.http import FileResponse
from django.conf import settings
import os
import pandas
from stock.models import UploadBarang,UploadBarangList,LogTransaksi
import uuid
import random
from posmimail import posmiMail
from pos.models import Penjualan,PenjualanDetail
from django.contrib.auth import authenticate
from cms.models import GantiEmail, Testimoni
from payment.views import cekKodeToko,cekKodeVoucher,cekLisensi,getAdmin
from payment.models import GaransiRefund
from django.contrib.auth import logout
from .forms import FormInputBarang


def bulannya(bulannya):
    if bulannya==1:
        return "Januari"
    elif bulannya==2:
        return "Februari"
    elif bulannya==3:
        return "Maret"
    elif bulannya==4:
        return "April"
    elif bulannya==5:
        return "Mei"
    elif bulannya==6:
        return "Juni"
    elif bulannya==7:
        return "Juli"
    elif bulannya==8:
        return "Agustus"
    elif bulannya==9:
        return "September"
    elif bulannya==10:
        return "Oktober"
    elif bulannya==11:
        return "November"
    elif bulannya==12:
        return "Desember"
    
def addLog(user,cabang,transaksi,keterangan):
    try:
        if(user):
            logtransaksi = LogTransaksi()
            logtransaksi.user=user
            logtransaksi.cabang=cabang
            logtransaksi.transaksi=transaksi
            logtransaksi.keterangan=keterangan
            logtransaksi.save()
        else:
            logtransaksi = LogTransaksi()
            logtransaksi.transaksi=transaksi
            logtransaksi.keterangan=keterangan
            logtransaksi.save()
        return True
    except:
        return False
    
def cekTotal(data):
    if data==None:
        return 0
    else:
        return data

def index(request):
    if request.user.is_authenticated:
        # is_staff=True = platform admin → /management/
        if request.user.is_staff:
            return HttpResponseRedirect('/management/')
        # Owner Korporasi (punya owner_profile, tidak punya userprofile) → /owner/
        if hasattr(request.user, 'owner_profile') and not hasattr(request.user, 'userprofile'):
            return HttpResponseRedirect('/owner/')
        if request.user.is_superuser:
            try:
                _ = request.user.userprofile
            except Exception:
                return HttpResponseRedirect('/management/')
            penjualan = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=datetime.datetime.now().month) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False))
            total_penjualan_sebulan = penjualan.aggregate(jumlah=Sum('total'))
            total_penjualan_setahun = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(jumlah=Sum('total'))
            bulan = bulannya(datetime.datetime.now().month)
            tahun = datetime.datetime.now().year
            pending_penjualan = Penjualan.objects.all().filter(Q(is_paid=False) & Q(cabang=request.user.userprofile.cabang) & Q(created_at__month=datetime.datetime.now().month) & Q(created_at__year=datetime.datetime.now().year)).aggregate(jumlah=Count('total'))
            jumlah_penjualan = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=datetime.datetime.now().month) & Q(updated_at__year=datetime.datetime.now().year)  & Q(is_void=False)).aggregate(jumlah=Count('total'))
        
            # mengisi grafik penjualan
            april = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=4) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            januari = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=1) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            februari = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=2) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            maret = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=3) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            mei = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=5) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            juni = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=6) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            juli = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=7) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            agustus = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=8) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            september = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=9) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            oktober = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=10) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            november = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=11) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            desember = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=12) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).aggregate(total=Sum('total'))
            
            # mengisi top 5 barang terjual
            
            jumlah_barang = Barang.objects.all().filter(cabang=request.user.userprofile.cabang).count()

            barang = Barang.objects.all().filter(cabang=request.user.userprofile.cabang).order_by('-jumlah_dibeli','updated_at')[:5]
            list_nama_barang = []
            list_jumlah_barang = []
            for bar in barang:
                list_nama_barang.append(bar.nama)
                list_jumlah_barang.append(bar.jumlah_dibeli)
            
            print(request.user.userprofile.cabang)
            # mengisi jumlah transaksi pengguna
            liga_kasir = Penjualan.objects.filter(
                Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) &
                Q(updated_at__month=datetime.datetime.now().month) &
                Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)
            ).values('user__username').annotate(jumlah=Count('user'))
            nama_kasir = []
            jumlah_transaksi = []
            for kasir in liga_kasir:
                nama_kasir.append(kasir['user__username'] or '-')
                jumlah_transaksi.append(kasir['jumlah'])
            
            print(liga_kasir)

            # mengisi jumlah pembayaran cash vs transfer
            liga_metode = Penjualan.objects.all().filter(Q(is_paid=True) & Q(cabang=request.user.userprofile.cabang) & Q(updated_at__month=datetime.datetime.now().month) & Q(updated_at__year=datetime.datetime.now().year) & Q(is_void=False)).values('metode').annotate(jumlah=Count('metode')).order_by()
            print(liga_metode)
            jumlah_metode = []
            # apakah metode cash?
            try:
                if(liga_metode[0]['metode']==0):
                    # tambahkan jumlah metodenya
                    jumlah_metode.append(liga_metode[0]['jumlah'])
                else:
                    # jika tidak ada maka kasih angka 0
                    jumlah_metode.append(0)
            except:
                jumlah_metode.append(0)
            # apakah metode transfer?
            try:
                if(liga_metode[1]['metode']==1):
                    jumlah_metode.append(liga_metode[1]['jumlah'])
                else:
                    jumlah_metode.append(0)
            except:
                jumlah_metode.append(0)
            
            print(jumlah_metode)
            cabang = request.user.userprofile.cabang
            refund_garansi = GaransiRefund.objects.filter(cabang=cabang).order_by('-created_at').first()
            refund_jumlah_transaksi = 0
            if refund_garansi:
                refund_jumlah_transaksi = Penjualan.objects.filter(
                    cabang=cabang,
                    is_paid=True,
                    is_void=False,
                    created_at__gte=refund_garansi.tanggal_aktivasi,
                    created_at__lte=datetime.datetime.now(),
                ).count()

            context = {
                'bulan':bulan,
                'tahun':tahun,
                'total_penjualan_sebulan':total_penjualan_sebulan['jumlah'],
                'total_penjualan_setahun':total_penjualan_setahun['jumlah'],
                'pending_penjualan':pending_penjualan['jumlah'],
                'jumlah_penjualan':jumlah_penjualan['jumlah'],
                'januari':cekTotal(januari['total']),
                'april':cekTotal(april['total']),
                'febuari':cekTotal(februari['total']),
                'maret':cekTotal(maret['total']),
                'mei':cekTotal(mei['total']),
                'juni':cekTotal(juni['total']),
                'juli':cekTotal(juli['total']),
                'agustus':cekTotal(agustus['total']),
                'september':cekTotal(september['total']),
                'oktober':cekTotal(oktober['total']),
                'november':cekTotal(november['total']),
                'desember':cekTotal(desember['total']),
                'list_nama_barang':list_nama_barang,
                'list_jumlah_barang':list_jumlah_barang,
                'nama_kasir':nama_kasir,
                'jumlah_transaksi':jumlah_transaksi,
                'jumlah_metode':jumlah_metode,
                'jumlah_barang':jumlah_barang,
                'refund_garansi':refund_garansi,
                'refund_jumlah_transaksi':refund_jumlah_transaksi,
                'refund_auto_limit':50
            }
            return render(request,'administrator/index.html',context)
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def verifikasiEmail(request, token):
    """Verifikasi email toko via link yang dikirim saat registrasi."""
    from cms.models import EmailVerificationToken
    try:
        ev = EmailVerificationToken.objects.get(token=token)
        if ev.is_valid:
            ev.cabang.is_email_verified = True
            ev.cabang.save(update_fields=['is_email_verified'])
            ev.delete()
            messages.success(request, "Email toko berhasil diverifikasi! Selamat menggunakan POSMI.")
        else:
            messages.error(request, "Link verifikasi sudah kadaluarsa. Minta kirim ulang dari dashboard.")
    except EmailVerificationToken.DoesNotExist:
        messages.error(request, "Link verifikasi tidak valid.")
    return HttpResponseRedirect('/cms/')


def kirimUlangVerifikasi(request):
    """Kirim ulang email verifikasi untuk toko yang belum diverifikasi."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')
    cabang = request.user.userprofile.cabang
    if cabang.is_email_verified:
        messages.add_message(request, messages.SUCCESS, "Email sudah terverifikasi.")
        return HttpResponseRedirect('/cms/')
    from cms.models import EmailVerificationToken
    import datetime as _dt
    EmailVerificationToken.objects.filter(cabang=cabang).delete()
    ev_token = EmailVerificationToken.objects.create(
        cabang=cabang,
        expired_at=_dt.datetime.now() + _dt.timedelta(hours=48)
    )
    verify_url = f"{request.scheme}://{request.get_host()}/cms/verifikasi-email/{ev_token.token}/"
    from posmimail import posmiMail
    posmiMail(
        "Verifikasi Email POSMI",
        f"Klik link berikut untuk verifikasi email toko Anda (48 jam):\n{verify_url}\n\n— Tim POSMI",
        address=cabang.email
    )
    messages.success(request, f"Email verifikasi dikirim ke {cabang.email}.")
    return HttpResponseRedirect('/cms/')


def infoToko(request):
    if request.user.is_authenticated:
        detailwallet=None
        if request.user.is_superuser:
            if(request.method=="POST"):
                cabang = Cabang.objects.get(id=request.user.userprofile.cabang.id)
                cabang.nama_toko = request.POST['nama_toko']
                cabang.nama_cabang = request.POST['nama_cabang']
                cabang.alamat_toko = request.POST['alamat_toko']
                cabang.telpon = request.POST['telpon']
                cabang.keterangan = request.POST['keterangan']
                cabang.save()
                
                messages.add_message(request,messages.SUCCESS,"Data Informasi Toko Berhasil Diubah.")
                return HttpResponseRedirect('/cms/')
            
            forminfotoko = FormInfoToko(instance=request.user.userprofile.cabang)
            cabang = request.user.userprofile.cabang
            detailwallet = DetailWalet.objects.all().filter(cabang=cabang)
            wallet = cabang.wallet
            kode_toko = cabang.prefix
            context = {
                'forms':forminfotoko,
                'wallet':wallet,
                'kode_toko':kode_toko,
                'detailwallet':detailwallet
            }
            return render(request,'administrator/components/info_toko.html',context)
        else:
                messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
                return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def daftarBarang(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            from payment.models import TokoAddon
            cabang = request.user.userprofile.cabang
            barangs = Barang.objects.all().filter(cabang=cabang)
            context = {
                'barangs': barangs,
                'jumlah_barang': barangs.count(),
                'addon_barcode_aktif': bool(TokoAddon.get_for_cabang(cabang, TokoAddon.ADDON_BARCODE)),
            }
            return render(request,'administrator/components/list_barang.html',context)
        else:
                messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
                return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def transaksiBulanBerjalan(request):
    """Transaksi bulan berjalan — tampilan proper dengan filter kasir & download."""
    if not request.user.is_authenticated:
        return HttpResponseRedirect('/login/')
    if not request.user.is_superuser:
        return HttpResponseRedirect('/')

    now    = datetime.datetime.now()
    cabang = request.user.userprofile.cabang

    # Bulan/tahun bisa dipilih (default = bulan berjalan)
    bulan_param = int(request.GET.get('bulan', now.month))
    tahun_param = int(request.GET.get('tahun', now.year))
    kasir_id    = request.GET.get('kasir', '')

    BULAN_LIST = ['','Januari','Februari','Maret','April','Mei','Juni',
                  'Juli','Agustus','September','Oktober','November','Desember']

    # Daftar kasir toko untuk filter
    from stock.models import UserProfile
    kasir_list = UserProfile.objects.filter(cabang=cabang, is_active=True).select_related('user')

    # Query utama
    qs = Penjualan.objects.filter(
        Q(is_paid=True) & Q(cabang=cabang) & Q(is_void=False) &
        Q(tgl_bayar__month=bulan_param) & Q(tgl_bayar__year=tahun_param)
    )
    if kasir_id:
        qs = qs.filter(user_id=kasir_id)
    qs = qs.select_related('user__userprofile').order_by('-tgl_bayar')

    total_penjualan = qs.aggregate(total=Sum('total'))['total'] or 0
    jumlah_transaksi = qs.count()

    # Tanggal bertransaksi (untuk rekap harian)
    tanggal_bertransaksi = list(
        qs.exclude(tgl_bayar=None).dates('tgl_bayar', 'day', order='ASC')
    )

    # Periode yang tersedia (untuk selector bulan/tahun)
    periode_list = (
        Penjualan.objects
        .filter(is_paid=True, cabang=cabang, tgl_bayar__isnull=False)
        .dates('tgl_bayar', 'month', order='DESC')
    )

    context = {
        'transaksi':             qs,
        'bulannya':              BULAN_LIST[bulan_param],
        'tahunnya':              tahun_param,
        'bulan_param':           bulan_param,
        'tahun_param':           tahun_param,
        'total_penjualan':       total_penjualan,
        'jumlah_transaksi':      jumlah_transaksi,
        'kasir_list':            kasir_list,
        'kasir_id':              kasir_id,
        'tanggal_bertransaksi':  tanggal_bertransaksi,
        'periode_list':          periode_list,
        'BULAN_LIST':            BULAN_LIST,
    }
    return render(request, 'administrator/components/transaksi_cms.html', context)

def transaksiBulanLain(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            cabang = request.user.userprofile.cabang

            # Periode yang punya transaksi — EXCLUDE bulan berjalan
            now_qos = datetime.datetime.now()
            periode_list = (
                Penjualan.objects
                .filter(is_paid=True, cabang=cabang, tgl_bayar__isnull=False)
                .exclude(tgl_bayar__month=now_qos.month, tgl_bayar__year=now_qos.year)
                .dates('tgl_bayar', 'month', order='DESC')
            )

            # Daftar kasir untuk filter
            from stock.models import UserProfile
            kasir_list = UserProfile.objects.filter(cabang=cabang, is_active=True).select_related('user')

            if request.method == "POST":
                bulan    = int(request.POST['bulan'])
                tahun    = int(request.POST['tahun'])
                kasir_id = request.POST.get('kasir', '')

                qs = Penjualan.objects.filter(
                    Q(is_paid=True) & Q(cabang=cabang) &
                    Q(tgl_bayar__month=bulan) & Q(tgl_bayar__year=tahun) &
                    Q(is_void=False)
                )
                if kasir_id:
                    qs = qs.filter(user_id=kasir_id)
                transaksi = qs.select_related('user__userprofile').order_by('-tgl_bayar')

                context = {
                    'transaksi':     transaksi,
                    'bulannya':      bulannya(bulan),
                    'tahunnya':      tahun,
                    'periode_list':  periode_list,
                    'bulan_dipilih': bulan,
                    'tahun_dipilih': tahun,
                    'kasir_list':    kasir_list,
                    'kasir_id':      kasir_id,
                }
                return render(request, 'administrator/components/history_bulan_lain.html', context)

            context = {'periode_list': periode_list}
            return render(request, 'administrator/components/filter_history.html', context)
        else:
            messages.add_message(request, messages.SUCCESS, "Anda tidak memiliki ijin.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request, messages.SUCCESS, "Silakan Login terlebih dahulu.")
        return HttpResponseRedirect('/login/')

def profilSaya(request):
    if not request.user.is_authenticated:
        messages.add_message(request, messages.SUCCESS, "Silakan Login terlebih dahulu.")
        return HttpResponseRedirect('/login/')
    if not request.user.is_superuser:
        messages.add_message(request, messages.SUCCESS, "Anda tidak memiliki ijin.")
        return HttpResponseRedirect('/')

    profile = request.user.userprofile

    if request.method == 'POST':
        aksi = request.POST.get('aksi', '')

        if aksi == 'profil':
            next_url = request.POST.get('next', '/cms/profil/')
            nama_lengkap = request.POST.get('nama_lengkap', '').strip()
            if nama_lengkap:
                profile.nama_lengkap = nama_lengkap
            if 'foto' in request.FILES:
                file = request.FILES['foto']
                namafile = str(uuid.uuid4()) + '.' + file.name.split('.')[-1]
                file_name = default_storage.save(f'foto_profile/{namafile}', file)
                profile.foto = 'foto_profile/' + namafile
            profile.save()
            messages.success(request, 'Profil berhasil diperbarui.')
            return HttpResponseRedirect(next_url)

        elif aksi == 'password':
            pw_lama = request.POST.get('pw_lama', '')
            pw_baru = request.POST.get('pw_baru', '')
            pw_konfirmasi = request.POST.get('pw_konfirmasi', '')
            if not request.user.check_password(pw_lama):
                messages.error(request, 'Password lama tidak sesuai.')
                return HttpResponseRedirect('/cms/profil/')
            elif len(pw_baru) < 6:
                messages.error(request, 'Password baru minimal 6 karakter.')
                return HttpResponseRedirect('/cms/profil/')
            elif pw_baru != pw_konfirmasi:
                messages.error(request, 'Konfirmasi password tidak cocok.')
                return HttpResponseRedirect('/cms/profil/')
            else:
                request.user.set_password(pw_baru)
                request.user.save()
                logout(request)
                messages.success(request, 'Password berhasil diubah. Silakan masuk kembali.')
                return HttpResponseRedirect('/login/')

    return render(request, 'administrator/components/profile.html', {'profile': profile})
    
def editBarang(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            if request.method=="POST":
                try:
                    id_barang = request.GET['id']
                    
                    try:
                        barang = Barang.objects.get(Q(id=id_barang) & Q(cabang=request.user.userprofile.cabang))
                    except:
                        return HttpResponse(f"<div style='margin-top:100px;text-align:center;font-style:italic;font-weight:bold;font-size:15px'>Barang tidak ada dalam data toko {request.user.userprofile.cabang.nama_toko}.<br>Update Informasi Barang Tidak Dapat Dilakukan.<br><br><input type='button' value='Tutup Halaman Ini' onclick='window.close()'></div>")
                    
                    barang.barcode=request.POST['barcode']
                    barang.nama=request.POST['nama']
                    barang.satuan=request.POST['satuan']
                    barang.stok = int(request.POST['stok'])
                    barang.harga_ecer=int(request.POST['harga_ecer'])
                    barang.harga_grosir=int(request.POST['harga_grosir'])
                    barang.harga_beli = int(request.POST['harga_beli'])
                    barang.keterangan = request.POST['keterangan']
                    barang.save()
                    addLog(request.user,request.user.userprofile.cabang,"edit barang",f"Edit Barang {barang.nama} ({barang.id})Berhasil")
                    messages.add_message(request,messages.SUCCESS,f"Update Informasi '{barang.nama}' Berhasil.")
                except Exception as ex:
                    print(ex)
                    addLog(request.user,request.user.userprofile.cabang,"edit barang",f"Edit Barang {barang.nama} ({barang.id}) Gagal")
                    messages.add_message(request,messages.SUCCESS,"Update Barang Gagal.")
            try:
                id_barang = request.GET['id']
                barang = Barang.objects.get(Q(id=id_barang) & Q(cabang=request.user.userprofile.cabang))
                form = FormBarang(instance=barang)
                context = {
                    'form':form,
                    'id_barang':id_barang
                }
                return render(request,'administrator/components/edit_barang.html',context)
            except Exception as ex:
                print(ex)
                return HttpResponseRedirect('/cms/')
        else:
                messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
                return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def tambahBarang(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            if request.method=="POST":
                # print(request.POST)
                # print(request.FILES)
                tanggal_upload=datetime.datetime.now()
                df = pandas.read_excel(request.FILES['file'])
                # print(df)
                list_informasi=[]
                uploadbarang = UploadBarang()
                uploadbarang.cabang = request.user.userprofile.cabang
                uploadbarang.user = request.user
                uploadbarang.save()
                for index,data in df.iterrows():
                    try:
                        barcode = int(data['barcode'])
                        nama = data['nama']
                        satuan = str(data['satuan']).upper()
                        try:
                            stok=int(data['stok'])
                        except Exception as ex:
                            # print(ex)
                            stok=0
                        try:
                            harga_ecer = int(data['harga_ecer'])
                        except Exception as ex:
                            # print(ex)
                            harga_ecer=0
                        try:
                            harga_grosir=int(data['harga_grosir'])
                        except Exception as ex:
                            # print(ex)
                            harga_grosir=0
                        try:
                            min_beli_grosir=int(data['min_beli_grosir'])
                        except Exception as ex:
                            # print(ex)
                            min_beli_grosir=0
                        try:
                            harga_beli=int(data['harga_beli'])
                        except Exception as ex:
                            # print(ex)
                            harga_beli=0
                        try:
                            keterangan = data['keterangan']
                        except:
                            keterangan=None

                        informasi = {
                            'barcode':barcode,
                            'nama':nama,
                            'satuan':satuan,
                            'stok':stok,
                            'harga_ecer':harga_ecer,
                            'harga_grosir':harga_grosir,
                            'harga_beli':harga_beli,
                            'min_beli_grosir':min_beli_grosir,
                            'keterangan':keterangan
                        }
                        # print(informasi)
                        list_informasi.append(informasi)
                        uploadbaranglist = UploadBarangList()
                        uploadbaranglist.upload_barang=uploadbarang
                        uploadbaranglist.barcode=barcode
                        uploadbaranglist.nama=nama
                        uploadbaranglist.satuan = satuan
                        uploadbaranglist.stok=stok
                        uploadbaranglist.harga_ecer=harga_ecer
                        uploadbaranglist.harga_grosir=harga_grosir
                        uploadbaranglist.min_beli_grosir=min_beli_grosir
                        uploadbaranglist.harga_beli=harga_beli
                        uploadbaranglist.keterangan=keterangan
                        uploadbaranglist.save()

                    except Exception as ex:
                        print(ex)
                        barcode = None
                    # if(barcode):
                    #     print(barcode)

                # print(header)
                jumlah_barang = len(list_informasi)
                messages.add_message(request,messages.SUCCESS,"Silakan Cek terlebih dahulu barang yang masuk daftar upload.")
                context = {
                    'tanggal_upload':tanggal_upload,
                    'total_barang':jumlah_barang,
                    'list_informasi':list_informasi,
                    'id_uploadbarang':str(uploadbarang.id_upload)
                }
                return render(request,'administrator/components/tambah_barang_list.html',context)
            context={}
            return render(request,'administrator/components/tambah_barang.html',context)
        else:
                messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
                return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def downloadTemplate(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            lokasi_template = os.path.join(settings.BASE_DIR,'static/template/template.xlsx')
            file = open(lokasi_template,'rb')
            response = FileResponse(file,as_attachment=True,filename="template_barang_upload.xlsx")
            return response
        else:
                messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
                return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def konfirmasiUpload(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            try:
                id = request.GET['id']
                uploadbarang = UploadBarang.objects.get(id_upload=id)
                uploadbaranglist = UploadBarangList.objects.all().filter(upload_barang=uploadbarang)
                
                for baranglist in uploadbaranglist:
                    try:
                        barang = Barang.objects.get(Q(cabang = request.user.userprofile.cabang) & Q(barcode=baranglist.barcode))
                        barang.satuan=baranglist.satuan
                        barang.stok=baranglist.stok
                        barang.harga_ecer = baranglist.harga_ecer
                        barang.harga_grosir = baranglist.harga_grosir
                        barang.min_beli_grosir = baranglist.min_beli_grosir
                        barang.harga_beli = baranglist.harga_beli
                        barang.keterangan = baranglist.keterangan
                        barang.save()
                    except Exception as ex:
                        print(ex)
                        barang = Barang()
                        barang.nama = baranglist.nama
                        barang.barcode = baranglist.barcode
                        barang.cabang=request.user.userprofile.cabang
                        barang.satuan=baranglist.satuan
                        barang.stok=baranglist.stok
                        barang.harga_ecer = baranglist.harga_ecer
                        barang.harga_grosir = baranglist.harga_grosir
                        barang.min_beli_grosir = baranglist.min_beli_grosir
                        barang.harga_beli = baranglist.harga_beli
                        barang.keterangan = baranglist.keterangan
                        barang.save()
                addLog(request.user,request.user.userprofile.cabang,"tambah barang",f"Menambahkan  {len(uploadbaranglist)} Barang Berhasil.")
                uploadbarang.delete()
                barangs = Barang.objects.all().filter(cabang=request.user.userprofile.cabang)
                context = {
                    'barangs':barangs,
                    'jumlah_barang':barangs.count()
                }
                messages.add_message(request,messages.SUCCESS,"Update data barang sudah berhasil. Silakan cek data barang.")
                return render(request,'administrator/components/list_barang.html',context)
            except Exception as ex:
                print(ex)
                addLog(request.user,request.user.userprofile.cabang,"tambah barang",f"Menambahkan   Barang Gagal.")
                messages.add_message(request,messages.SUCCESS,"Terjadi kesalahan upload barang, silakan coba lagi...")
                context={}
                return render(request,'administrator/components/tambah_barang.html',context)
        else:
                messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
                return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def hapusBarang(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            try:
                id=request.GET['id']
                
                try:
                    barang=Barang.objects.get(Q(cabang=request.user.userprofile.cabang) & Q(id=id))
                except:
                    messages.add_message(request,messages.SUCCESS,'Barang Gagal dihapus.')
                    return HttpResponseRedirect('/cms/')
                
                info = f' dengan nama {barang.nama} dan barcode [{barang.barcode}] '
                barang.delete()
                messages.add_message(request,messages.SUCCESS,f"Barang {info} Berhasil Dihapus.")
                addLog(request.user,request.user.userprofile.cabang,"hapus barang",f"Hapus Barang {barang.nama} ({barang.id}) Berhasil.")
            except Exception as ex:
                print(ex)
                addLog(request.user,request.user.userprofile.cabang,"hapus barang",f"Hapus Barang {barang.nama} ({barang.id}) Gagal.")
                messages.add_message(request,messages.SUCCESS,"Terjadi kesalahan hapus barang, barang yang sudah pernah dijual tidak dapat dihapus.")
            barangs = Barang.objects.all().filter(cabang=request.user.userprofile.cabang)
            return HttpResponseRedirect('/cms/')
        else:
                messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
                return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def downloadBarang(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            barcode=[]
            nama=[]
            satuan=[]
            stok=[]
            harga_ecer=[]
            harga_grosir=[]
            min_beli_grosir=[]
            harga_beli=[]
            keterangan=[]

            barangs = Barang.objects.all().filter(cabang=request.user.userprofile.cabang).order_by('nama')
            for barang in barangs:
                barcode.append(barang.barcode)
                nama.append(barang.nama)
                satuan.append(barang.satuan)
                stok.append(barang.stok)
                harga_beli.append(barang.harga_beli)
                harga_ecer.append(barang.harga_ecer)
                harga_grosir.append(barang.harga_grosir)
                min_beli_grosir.append(barang.min_beli_grosir)
                keterangan.append(barang.keterangan)

            df = pandas.DataFrame({
                'barcode':barcode,
                'nama':nama,
                'satuan':satuan,
                'stok':stok,
                'harga_beli':harga_beli,
                'harga_ecer':harga_ecer,
                'harga_grosir':harga_grosir,
                'min_beli_grosir':min_beli_grosir,
                'keterangan':keterangan
            })

            lokasi_file = os.path.join(settings.BASE_DIR,f'media/download/barang/{request.user.userprofile.cabang.token}.xlsx')

            df.to_excel(lokasi_file,index=False)

            file = open(lokasi_file,'rb')
            addLog(request.user,request.user.userprofile.cabang,"download barang",f"Download daftar Barang Berhasil.")
            response = FileResponse(file,as_attachment=True,filename=f"{request.user.userprofile.cabang.nama_toko}.xlsx")
            return response
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def viewLog(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            log_list = LogTransaksi.objects.all().filter(cabang=request.user.userprofile.cabang).order_by('-created_at')
            context = {
                'log_list':log_list
            }
            return render(request,'administrator/components/log.html',context)
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def daftarKasir(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            cabang = request.user.userprofile.cabang
            list_user = (
                User.objects.filter(userprofile__cabang=cabang)
                .select_related('userprofile')
                .order_by('username')
            )
            max_kasir = cabang.paket.max_user_login if cabang.paket else 0
            kasir_aktif = UserProfile.objects.filter(
                cabang=cabang, user__is_superuser=False, is_active=True
            ).count()
            context = {
                'list_user': list_user,
                'max_kasir': max_kasir,
                'kasir_aktif': kasir_aktif,
                'at_limit': max_kasir > 0 and kasir_aktif >= max_kasir,
                'nama_paket': cabang.paket.nama if cabang.paket else 'Paket Dasar',
            }
            return render(request,'administrator/components/list_user.html',context)
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def tambahKasir(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            if request.method=="POST":
                pass1 = request.POST['password1']
                pass2 = request.POST['password2']
                print(pass1)
                print(pass2)
                if pass1!=pass2:
                    messages.add_message(request,messages.SUCCESS,'Penambahan kasir gagal, password dan konfirmasinya tidak sama. Silakan ulangi kembali.')
                    return render(request,'administrator/components/tambah_user.html')
                else:
                    try:
                        nama_lengkap = request.POST['nama_lengkap']
                        id_cabang = request.user.userprofile.cabang.id
                        cabang = Cabang.objects.get(id=id_cabang)
                        jumlah_kasir = cabang.jumlah_kasir+1
                        userid = f"{cabang.prefix}{jumlah_kasir}"
                        print(userid)
                        try:
                            usernya = User()
                            usernya.username=userid
                            usernya.password=pass1
                            usernya.save()
                            usernya.set_password(pass1)
                            usernya.save()


                            cabang.jumlah_kasir=cabang.jumlah_kasir+1
                            cabang.save()
                        except:
                            jumlah_kasir = cabang.jumlah_kasir+2
                            userid = f"{cabang.prefix}{jumlah_kasir}"
                            usernya = User()
                            usernya.username=userid
                            usernya.password=pass1
                            usernya.save()
                            usernya.set_password(pass1)
                            usernya.save()

                            cabang.jumlah_kasir=cabang.jumlah_kasir+2
                            cabang.save()

                        # Kasir baru aktif hanya jika kuota belum penuh
                        max_kasir = cabang.paket.max_user_login if cabang.paket else 0
                        kasir_aktif = UserProfile.objects.filter(
                            cabang=cabang, user__is_superuser=False, is_active=True
                        ).count()
                        bisa_aktif = (max_kasir == 0) or (kasir_aktif < max_kasir)

                        userprofile = UserProfile()
                        userprofile.user=usernya
                        userprofile.cabang=cabang
                        userprofile.nama_lengkap=nama_lengkap
                        userprofile.is_active=bisa_aktif
                        userprofile.save()

                        status_msg = "sudah aktif dan bisa melakukan transaksi" if bisa_aktif else \
                            f"ditambahkan dalam kondisi NONAKTIF karena kuota kasir paket {cabang.paket.nama if cabang.paket else ''} sudah penuh ({kasir_aktif}/{max_kasir}). Aktifkan melalui menu Daftar Pengguna."
                        message = f"Selamat!\n\nUser {nama_lengkap} dengan username {usernya} dan password {pass1} berhasil ditambahkan.\n\nUser {status_msg}\n\nUntuk login bisa melakukan akses ke: https://posmi.pythonanywhere.com/login/ \n\nTerima kasih sudah memilih POSMI sebagai aplikasi untuk penjualan di toko Sobat. Apabila ada kendala segera hubungi tim POSMI.\n\n\nSalam,\n\nSuryo Adhy Chandra\n------------------\nCreator POSMI\n\n\nEmail: adhy.chandra@live.co.uk\nWhatsapp: +6281213270275\nTelegram: @suryo_adhy"
                        posmiMail("Penambahan User Kasir",message=message,address=cabang.email)
                        if not bisa_aktif:
                            messages.add_message(request, messages.SUCCESS,
                                f"Kasir {nama_lengkap} ({usernya.username}) berhasil ditambahkan "
                                f"namun dalam kondisi nonaktif karena kuota kasir sudah penuh. "
                                f"Aktifkan melalui Daftar Pengguna."
                            )
                        
                        return HttpResponseRedirect('/cms/kasir/')
                    except Exception as ex:
                        print(ex)
                        return HttpResponseRedirect('/cms/kasir/tambah/')
            else:
                return render(request,'administrator/components/tambah_user.html')
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def detailPenjualan(request):
    if request.user.is_authenticated:
        if request.user.userprofile.is_active:
            try:
                nota = request.GET['id']
                penjualan = Penjualan.objects.get(Q(nota=nota) & Q(cabang=request.user.userprofile.cabang))
                penjualan_detail = PenjualanDetail.objects.all().filter(penjualan=penjualan)
                context = {
                    'penjualan':penjualan,
                    'penjualan_detail':penjualan_detail
                }
                return render(request,'detail-histori-penjualan.html',context)
            except Exception as ex:
                print(ex)
                return HttpResponse("<center><h4 style='margin-top:200px;font-style:italic'>Maaf tidak memiliki akses.</h4></center>")
        else:
            messages.add_message(request,messages.SUCCESS,"Pengguna telah dinonaktifkan, silakan hubungi pemilik toko untuk konfirmasi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def konfirmasiVoid(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            try:
                id_nota = request.GET['id']
                penjualan = Penjualan.objects.get(Q(nota=id_nota) & Q(cabang=request.user.userprofile.cabang))
                penjualan_detail = PenjualanDetail.objects.all().filter(penjualan=penjualan)
                context = {
                    'callback':f'/cms/void/ok/?id={id_nota}',
                    'nota':id_nota,
                    'penjualan':penjualan,
                    'penjualan_detail':penjualan_detail
                }
                return render(request,'administrator/components/konfirmasi-void.html',context)
            except Exception as ex:
                print(ex)
                return HttpResponse("<center><h4 style='margin-top:200px;font-style:italic'>Maaf tidak memiliki akses.</h4></center>")
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def tidakVoid(request):
    return HttpResponse("/cms/")

def okeVoid(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            try:
                if request.method == "POST":
                    alasan = request.POST['alasan']
                    id_nota = request.GET['id']
                    penjualan = Penjualan.objects.get(Q(nota=id_nota) & Q(cabang=request.user.userprofile.cabang))
                    penjualan_detail = PenjualanDetail.objects.all().filter(penjualan=penjualan)
                    for jual in penjualan_detail:
                        barang = jual.barang
                        barang.stok += jual.jumlah
                        barang.jumlah_dibeli -= jual.jumlah
                        barang.save()
                    penjualan.is_void=True
                    penjualan.alasan_void=alasan
                    penjualan.save()
                    
                    messages.add_message(request,messages.SUCCESS,f'Nota dengan id: {id_nota} berhasil dibatalkan (void).')
                return HttpResponseRedirect('/cms/')
            except Exception as ex:
                print(ex)
                return HttpResponse("<center><h4 style='margin-top:200px;font-style:italic'>Maaf tidak memiliki akses.</h4></center>")
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def gantiEmail(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            if request.method=="POST":
                email_baru = request.POST['email_baru']
                password = request.POST['password']

                user = request.user
                if authenticate(username=user.username,password=password):
                    # cek email apakah pernah ada
                    try:
                        cabang = Cabang.objects.get(email=email_baru)
                        messages.add_message(request,messages.SUCCESS,"Email sudah pernah didaftarkan di POSMI, silakan menggunakan email lain.")
                        return HttpResponseRedirect('/cms/')
                    except:
                        gantiemail = GantiEmail()
                        gantiemail.cabang = request.user.userprofile.cabang
                        gantiemail.clicked = False
                        gantiemail.email_baru = email_baru
                        gantiemail.user = request.user
                        gantiemail.save()
                        addLog(request.user,request.user.userprofile.cabang,"Ganti Email Toko",f"Request ganti email toko {request.user.userprofile.cabang.nama_toko} menjadi {email_baru}")
                        message = f"Sobat {request.user.userprofile.nama_lengkap},\n\n\nSobat telah melakukan permintaan perubahan email pada {datetime.datetime.now().strftime('%d/%h/%Y')}.\n\nSilakan klik link berikut ini untuk melanjutkan perubahan email:\n\nhttps://posmi.pythonanywhere.com/cms/{gantiemail.id}/ \n\nSebagai catatan, untuk link ini hanya bisa diakses 1 kali saja dan akan expired dalam 1 jam.\n\nApabila ada kendala, segera hubungi kami. Terima kasih sudah mempercayakan aplikasi kasir menggunakan POSMI.\n\n\nSalam,\n\nSuryo Adhy Chandra\n------------------\nCreator POSMI\n\n\nEmail: adhy.chandra@live.co.uk\nWhatsapp: +6281213270275\nTelegram: @suryo_adhy"
                        posmiMail("PERUBAHAN EMAIL TOKO",message,email_baru)
                        messages.add_message(request,messages.SUCCESS,f"Permintaan perubahan email sudah berhasil. Silakan sobat cek email baru {email_baru} dan klik tautan (link) untuk konfirmasi perubahan email. Terima kasih. ")
                else:
                    messages.add_message(request,messages.SUCCESS,"Silakan ulangi kembali, password admin untuk toko tidak sesuai.")
            return HttpResponseRedirect('/cms/')
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def konfirmasiEmail(request,id):
    try:
        id_email = id
        gantiemail = GantiEmail.objects.get(Q(id=id_email) & Q(clicked=False) & Q(expired__gt=datetime.datetime.now()))
        gantiemail.clicked=True
        gantiemail.save()
        print('status ganti email sudah diupdate')

        cabang = gantiemail.cabang
        cabang.email = gantiemail.email_baru.lower()
        cabang.save()
        print('email cabang sudah diupdate')
        
        addLog(gantiemail.user,cabang,"Ganti Email Toko","Ganti Email Toko Berhasil.")
    except Exception as ex:
        print(ex)
    pesan = f"Penggantian email sudah dikonfirmasi, silakan cek untuk email di halaman admin. Terima kasih kepercayaan Sobat menggunakan POSMI."
    context = {
        'pesan':pesan
    }
    return render(request,'konfirmasi-perubahan-email.html',context)    


    
def tambahKuotaAdmin(request,id):
    asal = "/cms/"
    print(id)
    try:
        kode_toko = id    
        info_registrasi="Penambahan Kuota"
        list_kuota=[data for data in range(50,1500,50)]
        list_biaya = []
        try:
            cabang = Cabang.objects.get(prefix=kode_toko)
            
            info_registrasi="Penambahan Kuota"
            
            nama_admin = getAdmin(kode_toko)
            context = {
                'kode_toko':kode_toko,
                'cabang':cabang,
                'nama_admin':nama_admin,
                'info_registrasi':info_registrasi,
                'list_kuota':list_kuota,
                'asal':asal,
                'list_biaya':list_biaya,
                'tipe':"kuota",
                'wallet': cabang.wallet,
            }
            return render(request,'registrasi/cek_lisensi.html',context)
        except Exception as ex:
            print(ex)
            messages.add_message(request,messages.SUCCESS,"Kode Toko Tidak Ditemukan.")    
            return HttpResponseRedirect("/cms/")        
    except Exception as ex:
        print(ex)
        messages.add_message(request,messages.SUCCESS,"Kode Toko Tidak Ditemukan.")
        return HttpResponseRedirect("/cms/")
    

def upgradePaketAdmin(request,id):
    asal = '/cms/'
    print(id)
    try:
        kode_toko = id    
        info_registrasi="Upgrade ke Paket Bisnis Kecil atau Medium"
        tipe = "upgrade" 
        info_registrasi=""
        list_kuota=[data for data in range(50,1500,50)]
        list_biaya = []
        try:
            cabang = Cabang.objects.get(prefix=kode_toko)
            
            if cabang.paket== None:
                info_registrasi="Upgrade ke Paket Bisnis Kecil atau Medium"
                daftarpaket = DaftarPaket.objects.all().filter(nama__in=['Bisnis Kecil','Bisnis Medium'])
                sisa_uang=0
                for daftar in daftarpaket:
                    data = {
                            "id":daftar.paket,
                            'nama':daftar.nama,
                        }
                    biaya_list=[]
                    if int(daftar.harga_per_bulan-sisa_uang)>0:
                        biaya_data = {
                            'nama':'bulanan',
                            'info': '1 bulanan',
                            'data':int(daftar.harga_per_bulan-sisa_uang)
                        }
                        biaya_list.append(biaya_data)
                    if int(daftar.harga_per_tiga_bulan-sisa_uang)>0:
                        biaya_data = {
                            'nama':'tigabulanan',
                            'info':'3 bulanan',
                            'data':int(daftar.harga_per_tiga_bulan-sisa_uang)
                        }
                        biaya_list.append(biaya_data)
                    if int(daftar.harga_per_enam_bulan-sisa_uang)>0:
                        biaya_data = {
                            'nama':'enambulanan',
                            'info':'6 bulanan',
                            'data':int(daftar.harga_per_enam_bulan-sisa_uang)
                        }
                        biaya_list.append(biaya_data)
                    if int(daftar.harga_per_tahun-sisa_uang)>0:
                        biaya_data = {
                            'nama':'tahunan',
                            'info':'1 tahunan',
                            'data':int(daftar.harga_per_tahun-sisa_uang)
                        }
                        biaya_list.append(biaya_data)
                    if int(daftar.harga_per_dua_tahun-sisa_uang)>0:
                        biaya_data = {
                            'nama':'duatahunan',
                            'info':'2 tahunan',
                            'data':int(daftar.harga_per_dua_tahun-sisa_uang)
                        }
                        biaya_list.append(biaya_data)
                    data['biaya']=biaya_list
                    list_biaya.append(data)
                print(list_biaya)
            elif cabang.paket.nama=="Bisnis Kecil":
                    info_registrasi="Upgrade ke Paket Bisnis Medium"
                    daftarpaket = DaftarPaket.objects.all().filter(nama__in=['Bisnis Medium'])
                    sisa_hari = int((cabang.lisensi_expired-datetime.datetime.now()).days)
                    sisa_uang = 0
                    # cek apakah sisa hari masih bulanan?
                    sisa_uang = (cabang.paket.harga_per_tahun/365)*sisa_hari
                    for daftar in daftarpaket:
                        data = {
                                "id":str(daftar.paket),
                                'nama':daftar.nama,
                        }
                        
                        biaya_list=[]
                        if int(daftar.harga_per_bulan-sisa_uang)>0:
                            biaya_data = {
                                'nama':'bulanan',
                                'info': '1 bulanan',
                                'data':int(daftar.harga_per_bulan-sisa_uang)
                            }
                            biaya_list.append(biaya_data)
                        if int(daftar.harga_per_tiga_bulan-sisa_uang)>0:
                            biaya_data = {
                                'nama':'tigabulanan',
                                'info':'3 bulanan',
                                'data':int(daftar.harga_per_tiga_bulan-sisa_uang)
                            }
                            biaya_list.append(biaya_data)
                        if int(daftar.harga_per_enam_bulan-sisa_uang)>0:
                            biaya_data = {
                                'nama':'enambulanan',
                                'info':'6 bulanan',
                                'data':int(daftar.harga_per_enam_bulan-sisa_uang)
                            }
                            biaya_list.append(biaya_data)
                        if int(daftar.harga_per_tahun-sisa_uang)>0:
                            biaya_data = {
                                'nama':'tahunan',
                                'info':'1 tahunan',
                                'data':int(daftar.harga_per_tahun-sisa_uang)
                            }
                            biaya_list.append(biaya_data)
                        if int(daftar.harga_per_dua_tahun-sisa_uang)>0:
                            biaya_data = {
                                'nama':'duatahunan',
                                'info':'2 tahunan',
                                'data':int(daftar.harga_per_dua_tahun-sisa_uang)
                            }
                            biaya_list.append(biaya_data)
                        data['biaya']=biaya_list
                        print(data)
                        list_biaya.append(data)
            else:
                messages.add_message(request,messages.SUCCESS,f"Paket untuk {cabang.nama_toko} sudah Medium tidak bisa melakukan upgrade. Yang bisa dilakukan adalah menambah kuota transaksi penjualan. Terima kasih.")
                return HttpResponseRedirect('/')         
            nama_admin = getAdmin(kode_toko)
            context = {
                'kode_toko':kode_toko,
                'cabang':cabang,
                'nama_admin':nama_admin,
                'info_registrasi':info_registrasi,
                'tipe':tipe,
                'list_kuota':list_kuota,
                'asal':asal,
                'list_biaya':list_biaya,
                'wallet': cabang.wallet,
            }
            return render(request,'registrasi/cek_lisensi.html',context)
        except Exception as ex:
            print(ex)
            messages.add_message(request,messages.SUCCESS,"Kode Toko Tidak Ditemukan.")    
    except Exception as ex:
        print(ex)
        messages.add_message(request,messages.SUCCESS,"Kode Toko Tidak Ditemukan.")
        return HttpResponseRedirect("/")
    
def gantiPassword(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            if request.method=="POST":
                print(request.POST)
                

            user_id = request.GET['id']

            cabang = request.user.userprofile.cabang
            user = User.objects.get(id=user_id)

            try:
                user_identifier = int(user.username[-1])

                if user_identifier==1:
                    status_logout=1
                else:
                    status_logout=0
            except:
                # langsung jadikan 1
                status_logout=1

            try:
                userprofile = UserProfile.objects.get(Q(user=user) & Q(cabang=cabang))
                context = {
                    'username':user.username,
                    'nama_lengkap':userprofile.nama_lengkap,
                    'callback': f'/cms/pass/change/?logout={status_logout}&id={user_id}'
                }
                return render(request,'administrator/components/ganti-password.html',context)
            except:
                return HttpResponse("<center><div style='font-style:italic;margin-top:30px;font-weight:bold'>Anda tidak memiliki hak akses.</div></center>")
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def updateStatusKasir(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:    
            try:
                user_id = request.GET['id']
                status = int(request.GET['status'])
                cabang = request.user.userprofile.cabang
                user = User.objects.get(id=user_id)

                if user.userprofile.cabang == cabang:
                    try:
                        if not user.is_superuser:
                            if status == 1:
                                # Nonaktifkan
                                user.userprofile.is_active = False
                                user.userprofile.save()
                                addLog(request.user, cabang, "Status Pengguna", f"Status Pengguna {user.username} berhasil dinonaktifkan")
                                messages.add_message(request, messages.SUCCESS, f"Pengguna {user.username} berhasil dinonaktifkan.")
                            else:
                                # Aktifkan — cek kuota paket
                                max_kasir = cabang.paket.max_user_login if cabang.paket else 0
                                if max_kasir > 0:
                                    kasir_aktif = UserProfile.objects.filter(
                                        cabang=cabang, user__is_superuser=False, is_active=True
                                    ).count()
                                    if kasir_aktif >= max_kasir:
                                        messages.add_message(
                                            request, messages.ERROR,
                                            f"Tidak bisa mengaktifkan {user.username}. "
                                            f"Kuota kasir paket {cabang.paket.nama} sudah penuh "
                                            f"({kasir_aktif}/{max_kasir} kasir aktif). "
                                            f"Nonaktifkan kasir lain terlebih dahulu."
                                        )
                                        return HttpResponseRedirect('/cms/kasir/')
                                user.userprofile.is_active = True
                                user.userprofile.save()
                                addLog(request.user, cabang, "Status Pengguna", f"Status Pengguna {user.username} berhasil diaktifkan kembali")
                                messages.add_message(request, messages.SUCCESS, f"Pengguna {user.username} berhasil diaktifkan kembali.")
                    except Exception as ex:
                        print(ex)
            except Exception as ex:
                print(ex)
                pass
            return HttpResponseRedirect('/cms/kasir/')
        
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def detailPengguna(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:    
            try:
                user_id = request.GET['id']
                cabang = request.user.userprofile.cabang
                user = User.objects.get(id=user_id)

                if user.userprofile.cabang == cabang:
                    try:
                        context = {
                            'pengguna':user
                        }
                        return render(request,'administrator/components/detail_pengguna.html',context=context)
                    except Exception as ex:
                        print(ex)
                        messages.add_message(request,messages.SUCCESS,"Anda Tidak memiliki hak akses.")
            except Exception as ex:
                print(ex)
                pass
            return HttpResponseRedirect('/cms/kasir/')
        
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')

def updateFoto(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:    
            try:
                user_id = request.GET['id']
                cabang = request.user.userprofile.cabang
                user = User.objects.get(id=user_id)

                if user.userprofile.cabang == cabang:
                    try:
                        # simpan files
                        if request.method == "POST":
                            file = request.FILES['foto']
                            namafile = str(uuid.uuid4()) + "." + str(file.name).split('.')[-1]
                            lokasi = os.path.join(settings.BASE_DIR,f'media/foto_profile/{namafile}')
                            print(lokasi)
                            file_name = default_storage.save(f"foto_profile/{namafile}",file)

                            user.userprofile.foto = 'foto_profile/'+namafile
                            user.userprofile.save()
                            messages.add_message(request,messages.SUCCESS,'Penggantian foto sudah berhasil.')
                            addLog(request.user,cabang,"Ganti Foto",f"Ganti Foto Profile Pengguna {user.username} Berhasil.")
                            return HttpResponseRedirect(f'/cms/user/detail/?id={user_id}')
                    except Exception as ex:
                        print(ex)
                        addLog(request.user,cabang,"Ganti Foto",f"Ganti Foto Profile Pengguna {user.username} Gagal.")
                        messages.add_message(request,messages.SUCCESS,"Anda Tidak memiliki hak akses.")
            except Exception as ex:
                print(ex)
                pass
            return HttpResponseRedirect('/cms/kasir/')
        
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def updateNama(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:    
            try:
                user_id = request.GET['id']
                cabang = request.user.userprofile.cabang
                user = User.objects.get(id=user_id)

                if user.userprofile.cabang == cabang:
                    try:
                        # simpan files
                        if request.method == "POST":
                            pengguna = request.POST['pengguna']
                            user.userprofile.nama_lengkap = pengguna
                            user.userprofile.save()
                            messages.add_message(request,messages.SUCCESS,f'Penggantian nama pengguna {user.username} sudah berhasil.')
                            addLog(request.user,cabang,"Ganti Nama",f"Ganti Nama Pengguna {user.username} Berhasil")
                            return HttpResponseRedirect(f'/cms/user/detail/?id={user_id}')
                    except Exception as ex:
                        addLog(request.user,cabang,"Ganti Nama",f"Ganti Nama Pengguna {user.username} Gagal.")
                        messages.add_message(request,messages.SUCCESS,"Anda Tidak memiliki hak akses.")
            except Exception as ex:
                print(ex)
                pass
            return HttpResponseRedirect('/cms/kasir/')
        
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def updatePassword(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:    
            try:
                user_id = request.GET['id']
                cabang = request.user.userprofile.cabang
                user = User.objects.get(id=user_id)

                if user.userprofile.cabang == cabang:
                    try:
                        # simpan files
                        if request.method == "POST":
                            password1 = request.POST['password1']
                            password2 = request.POST['password2']
                            if(password1==password2):
                                user.set_password(password1)
                                user.save()
                                if int(user.username[-1])==1:
                                    logout(request)
                                    messages.add_message(request,messages.SUCCESS,'Penggantian Password sudah berhasil. Silakan login kembali menggunakan password baru.')
                                    return HttpResponseRedirect('/')
                                else:
                                    messages.add_message(request,messages.SUCCESS,f'Penggantian Password sudah berhasil untuk Pengguna {user.username}.')
                                try:
                                    addLog(request.user,cabang,"Ganti Password",f"Ganti Password Pengguna {user.username} Berhasil")
                                except:
                                    addLog(user,cabang,"Ganti Password",f"Ganti Password Pengguna {user.username} Berhasil")
                            else:
                                addLog(request.user,cabang,"Ganti Password",f"Ganti Password Pengguna {user.username} Gagal.")
                                messages.add_message(request,messages.SUCCESS,'Penggantian Password Tidak berhasil. Password Baru dan Konfirmasinya tidak sama.')
                            return HttpResponseRedirect(f'/cms/user/detail/?id={user_id}')
                    except Exception as ex:
                        print(ex)
                        messages.add_message(request,messages.SUCCESS,"Anda Tidak memiliki hak akses.")
            except Exception as ex:
                print(ex)
                pass
            return HttpResponseRedirect('/cms/kasir/')
        
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
def tambahBarangSatuan(request):
    if request.user.is_authenticated:
        if request.user.is_superuser: 
            if request.method=="POST":
                print(request.POST)
                form = FormInputBarang(data=request.POST)
                if form.is_valid():
                    try:
                        barang = Barang.objects.get(Q(cabang=request.user.userprofile.cabang) & Q(barcode=int(request.POST['barcode'])))
                        form = FormInputBarang(data=request.POST)
                        messages.add_message(request,messages.SUCCESS,f"Barang {request.POST['nama']} gagal ditambahkan. Silakan cek apakah barcode sudah pernah dipakai dan sesuaikan kembali barcodenya. Terima kasih.")

                    except Exception as ex:
                        print(ex)
                        barang = Barang()
                        barang.barcode = int(request.POST['barcode'])
                        barang.cabang = request.user.userprofile.cabang
                        barang.created_by= request.user
                        barang.harga_beli = request.POST['harga_beli']
                        barang.harga_ecer = request.POST['harga_ecer']
                        barang.harga_grosir = request.POST['harga_grosir']
                        barang.min_beli_grosir = request.POST['min_beli_grosir']
                        barang.keterangan = request.POST['keterangan']
                        barang.nama = request.POST['nama']
                        barang.satuan = request.POST['satuan']
                        barang.save()
                        messages.add_message(request,messages.SUCCESS,f"Barang {barang.nama} dengan barcode [{barang.barcode}] berhasil ditambahkan.")
                        form = FormInputBarang()    
                else:
                    form = FormInputBarang(data=request.POST)
                    messages.add_message(request,messages.SUCCESS,"Terjadi kesalahan data barang. Silakan ulangi kembali.")
            else:
                form = FormInputBarang()

            context = {
                'form':form
            }

            return render(request,'administrator/components/input_barang.html',context)
        else:
            messages.add_message(request,messages.SUCCESS,"Anda tidak memiliki ijin untuk mengkases halaman admin posmi.")
            return HttpResponseRedirect('/')
    else:
        messages.add_message(request,messages.SUCCESS,"Silakan Login terlebih dahulu untuk bisa mengakses halaman admin posmi.")
        return HttpResponseRedirect('/login/')
    
    

def cetakRekapCMS(request):
    """Download rekap Excel dari CMS — bulanan atau harian, support filter kasir."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    now        = datetime.datetime.now()
    cabang     = request.user.userprofile.cabang
    mode       = request.GET.get('mode', 'bulanan')
    bulan      = int(request.GET.get('bulan', now.month))
    tahun      = int(request.GET.get('tahun', now.year))
    kasir_id   = request.GET.get('kasir', '')

    BULAN_LIST = ['','Januari','Februari','Maret','April','Mei','Juni',
                  'Juli','Agustus','September','Oktober','November','Desember']

    qs = Penjualan.objects.filter(
        Q(is_paid=True) & Q(cabang=cabang) & Q(is_void=False) &
        Q(tgl_bayar__month=bulan) & Q(tgl_bayar__year=tahun)
    )
    if kasir_id:
        qs = qs.filter(user_id=kasir_id)

    if mode == 'harian':
        tgl_str = request.GET.get('tanggal', '')
        try:
            import datetime as _dt
            tgl = _dt.datetime.strptime(tgl_str, '%Y-%m-%d').date()
            qs  = qs.filter(tgl_bayar__date=tgl)
            judul = f"Rekap Harian {tgl.strftime('%d %B %Y')} — {cabang.nama_toko}"
            nama_file = f"rekap_harian_{tgl.strftime('%Y%m%d')}_{cabang.prefix}.xlsx"
        except ValueError:
            return HttpResponseRedirect('/cms/')
    else:
        qs = qs.order_by('tgl_bayar')
        judul = f"Rekap Bulanan {BULAN_LIST[bulan]} {tahun} — {cabang.nama_toko}"
        nama_file = f"rekap_{tahun}{bulan:02d}_{cabang.prefix}.xlsx"

    qs = qs.select_related('user__userprofile').order_by('tgl_bayar')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap"
    hd_fill = PatternFill(start_color="1E2C5A", end_color="1E2C5A", fill_type="solid")
    hd_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin = Border(
        left=Side(style='thin', color='D8E3F5'), right=Side(style='thin', color='D8E3F5'),
        top=Side(style='thin', color='D8E3F5'),  bottom=Side(style='thin', color='D8E3F5'),
    )
    ws.merge_cells('A1:H1'); ws['A1'] = judul
    ws['A1'].font = Font(bold=True, size=13, color="1E2C5A")
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Dicetak: {now.strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(size=10, color="8899BB")
    ws.append([])
    headers = ['No. Nota', 'Tanggal', 'Waktu', 'Kasir', 'Customer', 'Metode', 'Total', 'Items']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = hd_font; cell.fill = hd_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin
    col_w = [14,14,8,20,18,10,16,7]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A5'

    total_grand = 0
    for idx, t in enumerate(qs):
        r = 5 + idx
        fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
        kasir_nama = ''
        try: kasir_nama = t.user.userprofile.nama_lengkap
        except: kasir_nama = t.user.username if t.user else '-'
        data = [t.no_nota, t.tgl_bayar.strftime('%d/%m/%Y') if t.tgl_bayar else '-',
                t.tgl_bayar.strftime('%H:%M') if t.tgl_bayar else '',
                kasir_nama, t.customer or '-',
                'Cash' if t.metode == 0 else 'Transfer', int(t.total), t.items]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin
            if fill.fill_type: cell.fill = fill
            if col == 7: cell.alignment = Alignment(horizontal='right')
        total_grand += int(t.total)

    sub_fill = PatternFill(start_color="F0F4FB", end_color="F0F4FB", fill_type="solid")
    fr = 5 + qs.count()
    ws.cell(row=fr, column=6, value="TOTAL").font = Font(bold=True, color="1E2C5A")
    tc = ws.cell(row=fr, column=7, value=total_grand)
    tc.font = Font(bold=True, color="1E2C5A"); tc.alignment = Alignment(horizontal='right')
    for col in range(1, 9): ws.cell(row=fr, column=col).fill = sub_fill; ws.cell(row=fr, column=col).border = thin

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nama_file}"'
    wb.save(response)
    return response


# ─── Add-ons Status di CMS ────────────────────────────────────────────────────

def cmsAddonStatus(request):
    """Tampilkan status add-on untuk toko ini."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    from payment.models import TokoAddon

    # Ambil semua addon untuk cabang ini (individual + via owner korporasi)
    addons_individual = list(TokoAddon.objects.filter(cabang=cabang))
    addons_korporasi  = list(TokoAddon.objects.filter(owner=cabang.owner)) if cabang.owner_id else []

    # Gabungkan: individual override korporasi
    addon_map = {}
    for a in addons_korporasi:
        addon_map[a.addon_type] = {'addon': a, 'via': 'korporasi'}
    for a in addons_individual:
        addon_map[a.addon_type] = {'addon': a, 'via': 'individual'}

    # Semua tipe addon
    all_types = [
        (TokoAddon.ADDON_BARCODE,  'Cetak Label Barcode',         'fa-barcode',   '#d97706', 50_000),
        (TokoAddon.ADDON_NOTA,     'Custom Template Nota',        'fa-file-text', '#7c3aed', 75_000),
        (TokoAddon.ADDON_AKUNTING, 'Laporan Akunting Otomatis',   'fa-calculator','#059669', 100_000),
    ]

    addon_display = []
    for atype, nama, icon, color, harga in all_types:
        entry = addon_map.get(atype)
        addon_display.append({
            'type': atype,
            'nama': nama,
            'icon': icon,
            'color': color,
            'harga': harga,
            'addon': entry['addon'] if entry else None,
            'via': entry['via'] if entry else None,
            'aktif': entry['addon'].is_active if entry else False,
        })

    context = {
        'cabang': cabang,
        'addon_display': addon_display,
    }
    return render(request, 'administrator/components/addon_status.html', context)


def cmsAddonAktifkan(request, addon_type):
    """Halaman registrasi/aktivasi add-on dari CMS."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    from payment.models import TokoAddon

    ALL_ADDONS = {
        TokoAddon.ADDON_BARCODE: {
            'nama': 'Cetak Label Barcode',
            'icon': 'fa-barcode',
            'color': '#d97706',
            'harga': 50_000,
            'deskripsi': 'Cetak label barcode produk langsung dari aplikasi POS. Mendukung berbagai ukuran label dan printer thermal.',
        },
        TokoAddon.ADDON_NOTA: {
            'nama': 'Custom Template Nota',
            'icon': 'fa-file-text',
            'color': '#7c3aed',
            'harga': 75_000,
            'deskripsi': 'Buat template struk/nota sesuai kebutuhan toko: tambahkan logo, pesan promosi, dan tata letak kustom.',
        },
        TokoAddon.ADDON_AKUNTING: {
            'nama': 'Laporan Akunting Otomatis',
            'icon': 'fa-calculator',
            'color': '#059669',
            'harga': 100_000,
            'deskripsi': 'Laporan keuangan otomatis meliputi laba rugi, arus kas, dan neraca berdasarkan data transaksi POS.',
        },
    }

    if addon_type not in ALL_ADDONS:
        return render(request, 'administrator/components/addon_status.html', {'cabang': cabang, 'addon_display': []})

    addon_info = ALL_ADDONS[addon_type]

    if request.method == 'POST':
        harga = addon_info['harga']
        expired_at = cabang.lisensi_expired

        now = datetime.datetime.now()
        addon, created = TokoAddon.objects.get_or_create(
            cabang=cabang, addon_type=addon_type,
            defaults={
                'expired_at': expired_at,
                'harga_dibayar': harga,
                'status': TokoAddon.STATUS_AKTIF,
                'activated_at': now,
            }
        )
        if not created:
            addon.status = TokoAddon.STATUS_AKTIF
            addon.activated_at = now
            addon.expired_at = expired_at
            addon.harga_dibayar = harga
            addon.save(update_fields=['status', 'activated_at', 'expired_at', 'harga_dibayar', 'updated_at'])

        posmiMail(
            f"ADD-ON AKTIF: {addon_info['nama']}",
            f"Toko: {cabang.nama_toko} ({cabang.prefix})\n"
            f"Add-on: {addon_info['nama']}\n"
            f"Harga: Rp {harga:,}\n"
            f"Aktif sejak: {now.strftime('%d/%m/%Y %H:%M')}\n"
            f"Berlaku s.d.: {expired_at}",
            address="adhy.chandra@live.co.uk"
        )

        # Render kembali halaman addon status setelah submit
        addons_individual = list(TokoAddon.objects.filter(cabang=cabang))
        addons_korporasi  = list(TokoAddon.objects.filter(owner=cabang.owner)) if cabang.owner_id else []
        addon_map = {}
        for a in addons_korporasi:
            addon_map[a.addon_type] = {'addon': a, 'via': 'korporasi'}
        for a in addons_individual:
            addon_map[a.addon_type] = {'addon': a, 'via': 'individual'}
        all_types = [
            (TokoAddon.ADDON_BARCODE,  'Cetak Label Barcode',        'fa-barcode',   '#d97706', 50_000),
            (TokoAddon.ADDON_NOTA,     'Custom Template Nota',       'fa-file-text', '#7c3aed', 75_000),
            (TokoAddon.ADDON_AKUNTING, 'Laporan Akunting Otomatis',  'fa-calculator','#059669', 100_000),
        ]
        addon_display = []
        for atype, nama, icon, color, harga_item in all_types:
            entry = addon_map.get(atype)
            addon_display.append({
                'type': atype, 'nama': nama, 'icon': icon, 'color': color, 'harga': harga_item,
                'addon': entry['addon'] if entry else None,
                'via': entry['via'] if entry else None,
                'aktif': entry['addon'].is_active if entry else False,
            })
        return render(request, 'administrator/components/addon_status.html', {
            'cabang': cabang,
            'addon_display': addon_display,
        })

    existing = TokoAddon.objects.filter(cabang=cabang, addon_type=addon_type).first()

    context = {
        'cabang': cabang,
        'addon_type': addon_type,
        'addon': addon_info,
        'existing': existing,
    }
    return render(request, 'administrator/components/addon_aktifkan.html', context)


def _get_laporan_data(cabang, bulan, tahun):
    """Hitung data laporan akunting untuk satu bulan."""
    from pos.models import PenjualanDetail
    from django.db.models import F

    qs = Penjualan.objects.filter(
        cabang=cabang, is_paid=True, is_void=False,
        tgl_bayar__month=bulan, tgl_bayar__year=tahun,
    )
    total_pendapatan = qs.aggregate(t=Sum('total'))['t'] or 0
    jumlah_transaksi = qs.count()

    details = PenjualanDetail.objects.filter(penjualan__in=qs)
    hpp = details.aggregate(t=Sum(F('jumlah') * F('barang__harga_beli')))['t'] or 0
    laba_kotor = int(total_pendapatan) - int(hpp)

    cash     = qs.filter(metode=0).aggregate(t=Sum('total'))['t'] or 0
    transfer = qs.filter(metode=1).aggregate(t=Sum('total'))['t'] or 0

    top_produk = list(
        details.values('barang__nama')
               .annotate(qty=Sum('jumlah'), nilai=Sum('total'))
               .order_by('-qty')[:10]
    )
    return {
        'total_pendapatan': int(total_pendapatan),
        'jumlah_transaksi': jumlah_transaksi,
        'hpp': int(hpp),
        'laba_kotor': laba_kotor,
        'cash': int(cash),
        'transfer': int(transfer),
        'top_produk': top_produk,
    }


def cmsLaporanAkunting(request):
    """Halaman utama laporan akunting — pilih bulan/tahun."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')
    cabang = request.user.userprofile.cabang
    from payment.models import TokoAddon
    if not TokoAddon.get_for_cabang(cabang, TokoAddon.ADDON_AKUNTING):
        return HttpResponseRedirect('/cms/addons/')

    available = (
        Penjualan.objects
        .filter(cabang=cabang, is_paid=True, is_void=False)
        .dates('tgl_bayar', 'month', order='DESC')
    )
    bulan_list = [{'bulan': d.month, 'tahun': d.year} for d in available]
    years = sorted({d['tahun'] for d in bulan_list}, reverse=True)
    current_year = datetime.datetime.now().year
    default_year = current_year if current_year in years else (years[0] if years else current_year)
    context = {
        'cabang': cabang,
        'bulan_list': bulan_list,
        'available_years': years,
        'default_year': default_year,
    }
    return render(request, 'administrator/components/laporan_akunting.html', context)


def cmsLaporanAkuntingPreview(request):
    """Fragment preview laporan untuk bulan/tahun terpilih."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')
    cabang = request.user.userprofile.cabang
    try:
        bulan = int(request.GET.get('bulan', 0))
        tahun = int(request.GET.get('tahun', 0))
    except (ValueError, TypeError):
        return HttpResponse('')
    if not bulan or not tahun:
        return HttpResponse('')

    data = _get_laporan_data(cabang, bulan, tahun)
    context = {'cabang': cabang, 'bulan': bulan, 'tahun': tahun, **data}
    return render(request, 'administrator/components/laporan_akunting_preview.html', context)


def cmsLaporanAkuntingExcel(request):
    """Download laporan akunting sebagai file Excel."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')
    cabang = request.user.userprofile.cabang
    try:
        bulan = int(request.GET.get('bulan', 0))
        tahun = int(request.GET.get('tahun', 0))
    except (ValueError, TypeError):
        return HttpResponseRedirect('/cms/laporan-akunting/')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse as HR

    NAMA_BULAN = ['','Januari','Februari','Maret','April','Mei','Juni',
                  'Juli','Agustus','September','Oktober','November','Desember']

    data  = _get_laporan_data(cabang, bulan, tahun)
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = f'Laporan {NAMA_BULAN[bulan]} {tahun}'

    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = f'LAPORAN AKUNTING — {cabang.nama_toko}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Periode: {NAMA_BULAN[bulan]} {tahun}'
    ws['A2'].alignment = Alignment(horizontal='center')

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1e2c5a')

    def hdr(cell, val):
        cell.value = val
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    def cell(c, val, bold=False, fmt=None):
        c.value = val
        if bold: c.font = Font(bold=True)
        if fmt:  c.number_format = fmt
        c.border = border

    # Ringkasan
    ws.append([])
    ws.append(['RINGKASAN KEUANGAN'])
    ws['A4'].font = Font(bold=True, size=12)
    ws.append(['Keterangan', 'Nilai'])
    hdr(ws['A5'], 'Keterangan'); hdr(ws['B5'], 'Nilai')

    rows = [
        ('Total Pendapatan', data['total_pendapatan']),
        ('HPP (Harga Pokok Penjualan)', data['hpp']),
        ('Laba Kotor', data['laba_kotor']),
        ('Jumlah Transaksi', data['jumlah_transaksi']),
        ('Tunai (Cash)', data['cash']),
        ('Transfer', data['transfer']),
    ]
    for r, (label, val) in enumerate(rows, start=6):
        cell(ws.cell(r, 1), label, bold=True)
        cell(ws.cell(r, 2), val, fmt='#,##0')

    # Top produk
    ws.append([])
    r0 = ws.max_row + 1
    ws.cell(r0, 1).value = 'TOP PRODUK TERJUAL'
    ws.cell(r0, 1).font = Font(bold=True, size=12)
    r0 += 1
    hdr(ws.cell(r0, 1), 'Nama Produk')
    hdr(ws.cell(r0, 2), 'Qty Terjual')
    hdr(ws.cell(r0, 3), 'Total Nilai')
    for p in data['top_produk']:
        r0 += 1
        cell(ws.cell(r0, 1), p['barang__nama'])
        cell(ws.cell(r0, 2), p['qty'], fmt='#,##0')
        cell(ws.cell(r0, 3), p['nilai'], fmt='#,##0')

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    resp = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="laporan_akunting_{NAMA_BULAN[bulan]}_{tahun}.xlsx"'
    wb.save(resp)
    return resp


def cmsLaporanAkuntingCetak(request):
    """Halaman cetak/PDF laporan akunting."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')
    cabang = request.user.userprofile.cabang
    try:
        bulan = int(request.GET.get('bulan', 0))
        tahun = int(request.GET.get('tahun', 0))
    except (ValueError, TypeError):
        return HttpResponseRedirect('/cms/laporan-akunting/')
    data  = _get_laporan_data(cabang, bulan, tahun)
    context = {'cabang': cabang, 'bulan': bulan, 'tahun': tahun, **data}
    return render(request, 'administrator/components/laporan_akunting_cetak.html', context)


def cmsCetakBarcodePage(request):
    """Halaman pilih produk untuk cetak barcode (HTMX fragment)."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    from payment.models import TokoAddon, AddonConfig

    if not TokoAddon.get_for_cabang(cabang, TokoAddon.ADDON_BARCODE):
        return HttpResponseRedirect('/cms/addons/')

    q = request.GET.get('q', '').strip()
    barangs = Barang.objects.filter(cabang=cabang).order_by('nama')
    if q:
        barangs = barangs.filter(nama__icontains=q)

    cfg_obj = AddonConfig.objects.filter(cabang=cabang, addon_type=TokoAddon.ADDON_BARCODE).first()
    barcode_config = cfg_obj.config if cfg_obj else {}

    context = {
        'barangs': barangs,
        'q': q,
        'barcode_config': barcode_config,
        'cabang': cabang,
    }
    return render(request, 'administrator/components/cetak_barcode_list.html', context)


def cmsCetakBarcode(request, barang_id):
    """Halaman cetak label barcode untuk satu produk."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    from payment.models import TokoAddon, AddonConfig

    addon = TokoAddon.get_for_cabang(cabang, TokoAddon.ADDON_BARCODE)
    if not addon:
        return HttpResponseRedirect('/cms/barang/daftar/')

    barang = Barang.objects.filter(id=barang_id, cabang=cabang).first()
    if not barang:
        return HttpResponseRedirect('/cms/barang/daftar/')

    cfg_obj = AddonConfig.objects.filter(cabang=cabang, addon_type=TokoAddon.ADDON_BARCODE).first()
    barcode_config = cfg_obj.config if cfg_obj else {}

    jumlah = int(request.GET.get('jumlah', 1))

    context = {
        'barang': barang,
        'cabang': cabang,
        'barcode_config': barcode_config,
        'jumlah': jumlah,
        'kolom': barcode_config.get('kolom_per_baris', 2),
    }
    return render(request, 'administrator/components/cetak_barcode.html', context)


def cmsAddonSettings(request, addon_type):
    """Halaman pengaturan add-on yang sudah aktif."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    from payment.models import TokoAddon, AddonConfig

    valid_types = [c[0] for c in TokoAddon.ADDON_CHOICES]
    if addon_type not in valid_types:
        return HttpResponseRedirect('/cms/addons/')

    addon_obj = TokoAddon.objects.filter(cabang=cabang, addon_type=addon_type).first()
    if not addon_obj or not addon_obj.is_active:
        return HttpResponseRedirect('/cms/addons/')

    config_obj, _ = AddonConfig.objects.get_or_create(cabang=cabang, addon_type=addon_type)

    # Migrasi field lama: footer_text → footer_text_1
    if addon_type == TokoAddon.ADDON_NOTA:
        cfg = config_obj.config
        if cfg.get('footer_text') and not cfg.get('footer_text_1'):
            cfg['footer_text_1'] = cfg.pop('footer_text')
            config_obj.config = cfg
            config_obj.save(update_fields=['config'])
        if cfg.get('header_text') and not cfg.get('sub_header'):
            cfg['sub_header'] = cfg.pop('header_text')
            config_obj.config = cfg
            config_obj.save(update_fields=['config'])

    if request.method == 'POST':
        new_config = {}
        if addon_type == TokoAddon.ADDON_BARCODE:
            new_config = {
                'ukuran_label':      request.POST.get('ukuran_label', '38x25'),
                'tipe_barcode':      request.POST.get('tipe_barcode', 'CODE128'),
                'kolom_per_baris':   int(request.POST.get('kolom_per_baris', 2)),
                'tampilkan_harga':   request.POST.get('tampilkan_harga') == 'on',
                'tampilkan_nama_toko': request.POST.get('tampilkan_nama_toko') == 'on',
                'tampilkan_nama_produk': request.POST.get('tampilkan_nama_produk') == 'on',
            }
        elif addon_type == TokoAddon.ADDON_NOTA:
            new_config = {
                'main_header':      request.POST.get('main_header', '').strip(),
                'sub_header':       request.POST.get('sub_header', '').strip(),
                'footer_text_1':    request.POST.get('footer_text_1', '').strip(),
                'footer_text_2':    request.POST.get('footer_text_2', '').strip(),
                'lebar_nota':       request.POST.get('lebar_nota', '58mm'),
                'tampilkan_no_nota': request.POST.get('tampilkan_no_nota') == 'on',
                'tampilkan_tanggal': request.POST.get('tampilkan_tanggal') == 'on',
                'tampilkan_kasir':  request.POST.get('tampilkan_kasir') == 'on',
            }
        elif addon_type == TokoAddon.ADDON_AKUNTING:
            new_config = {
                'email_laporan':    request.POST.get('email_laporan', '').strip(),
                'periode_default':  request.POST.get('periode_default', 'bulanan'),
                'kirim_otomatis':   request.POST.get('kirim_otomatis') == 'on',
            }
        config_obj.config = new_config
        config_obj.save()

        context = {
            'cabang': cabang,
            'addon_type': addon_type,
            'addon_obj': addon_obj,
            'config': config_obj,
            'saved': True,
        }
        return render(request, 'administrator/components/addon_settings.html', context)

    context = {
        'cabang': cabang,
        'addon_type': addon_type,
        'addon_obj': addon_obj,
        'config': config_obj,
        'saved': False,
    }
    return render(request, 'administrator/components/addon_settings.html', context)


# ─── Laporan Perubahan Stok Per Barang ────────────────────────────────────────

def laporanStokBarang(request):
    """Tampilkan riwayat perubahan stok untuk satu barang (via HTMX)."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang   = request.user.userprofile.cabang
    barangs  = Barang.objects.filter(cabang=cabang).order_by('nama')
    barang_id = request.GET.get('barang_id') or request.POST.get('barang_id')
    barang_sel = None
    history = []

    if barang_id:
        try:
            barang_sel = Barang.objects.get(id=barang_id, cabang=cabang)
            # 1. Penjualan (pengurangan stok)
            from pos.models import PenjualanDetail
            penjualan_details = (
                PenjualanDetail.objects
                .filter(barang=barang_sel, penjualan__is_paid=True, penjualan__is_void=False)
                .select_related('penjualan')
                .order_by('-penjualan__tgl_bayar')
            )
            for pd in penjualan_details:
                history.append({
                    'tanggal': pd.penjualan.tgl_bayar,
                    'jenis': 'Penjualan',
                    'referensi': f"Nota {pd.penjualan.no_nota}",
                    'qty': -pd.jumlah,
                    'warna': '#dc3545',
                })

            # 2. Penerimaan dari Gudang (penambahan stok)
            if cabang.owner_id:
                from owner.models import PengirimanGudangItem, TransferStok
                items_gudang = PengirimanGudangItem.objects.filter(
                    barang_tujuan=barang_sel,
                    status='diterima',
                ).select_related('pengiriman').order_by('-pengiriman__created_at')
                for ig in items_gudang:
                    history.append({
                        'tanggal': ig.pengiriman.created_at,
                        'jenis': 'Terima dari Gudang',
                        'referensi': ig.pengiriman.nomor_pengiriman,
                        'qty': +ig.jumlah_dikirim,
                        'warna': '#059669',
                    })

                # 3. Transfer masuk (penambahan)
                transfer_masuk = TransferStok.objects.filter(
                    barang_tujuan=barang_sel, status='approved'
                ).order_by('-updated_at')
                for t in transfer_masuk:
                    history.append({
                        'tanggal': t.updated_at,
                        'jenis': f'Transfer dari {t.cabang_asal.nama_toko}',
                        'referensi': t.nomor_faktur or '-',
                        'qty': +t.jumlah,
                        'warna': '#4e73df',
                    })

                # 4. Transfer keluar (pengurangan)
                transfer_keluar = TransferStok.objects.filter(
                    barang_asal=barang_sel, status='approved'
                ).order_by('-updated_at')
                for t in transfer_keluar:
                    history.append({
                        'tanggal': t.updated_at,
                        'jenis': f'Transfer ke {t.cabang_tujuan.nama_toko}',
                        'referensi': t.nomor_faktur or '-',
                        'qty': -t.jumlah,
                        'warna': '#f59e0b',
                    })

            # Sort by tanggal desc
            history.sort(key=lambda x: x['tanggal'] if x['tanggal'] else __import__('datetime').datetime.min, reverse=True)

        except Barang.DoesNotExist:
            barang_sel = None

    context = {
        'cabang': cabang, 'barangs': barangs,
        'barang_sel': barang_sel, 'history': history,
    }
    return render(request, 'administrator/components/laporan_stok.html', context)


def downloadLaporanStok(request):
    """Download laporan perubahan stok per barang sebagai Excel."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang    = request.user.userprofile.cabang
    barang_id = request.GET.get('barang_id')

    try:
        barang_sel = Barang.objects.get(id=barang_id, cabang=cabang)
    except Barang.DoesNotExist:
        messages.add_message(request, messages.SUCCESS, "Barang tidak ditemukan.")
        return HttpResponseRedirect('/cms/')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Perubahan Stok"

    # Header info
    ws['A1'] = f"Laporan Perubahan Stok: {barang_sel.nama}"
    ws['A1'].font = Font(bold=True, size=13, color="1E2C5A")
    ws['A2'] = f"Barcode: {barang_sel.barcode}  |  Satuan: {barang_sel.satuan}  |  Stok Saat Ini: {barang_sel.stok}"
    ws['A2'].font = Font(size=11, color="8899BB")
    ws.append([])

    # Column headers
    headers = ['Tanggal', 'Jenis Perubahan', 'Referensi', 'Qty (+/-)']
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1E2C5A", end_color="1E2C5A", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12

    # Get history (same logic as laporanStokBarang)
    req_mock = type('r', (), {'user': request.user, 'GET': {'barang_id': barang_id}})()
    # Reuse logic by calling the view function's data gathering
    history = []
    from pos.models import PenjualanDetail
    for pd in PenjualanDetail.objects.filter(
        barang=barang_sel, penjualan__is_paid=True, penjualan__is_void=False
    ).select_related('penjualan').order_by('-penjualan__tgl_bayar'):
        history.append((pd.penjualan.tgl_bayar, 'Penjualan', f"Nota {pd.penjualan.no_nota}", -pd.jumlah))

    if cabang.owner_id:
        from owner.models import PengirimanGudangItem, TransferStok
        for ig in PengirimanGudangItem.objects.filter(barang_tujuan=barang_sel, status='diterima').select_related('pengiriman'):
            history.append((ig.pengiriman.created_at, 'Terima dari Gudang', ig.pengiriman.nomor_pengiriman, +ig.jumlah_dikirim))
        for t in TransferStok.objects.filter(barang_tujuan=barang_sel, status='approved'):
            history.append((t.updated_at, f'Transfer dari {t.cabang_asal.nama_toko}', t.nomor_faktur or '-', +t.jumlah))
        for t in TransferStok.objects.filter(barang_asal=barang_sel, status='approved'):
            history.append((t.updated_at, f'Transfer ke {t.cabang_tujuan.nama_toko}', t.nomor_faktur or '-', -t.jumlah))

    import datetime as _dt
    history.sort(key=lambda x: x[0] if x[0] else _dt.datetime.min, reverse=True)

    plus_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    minus_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row in history:
        r = ws.max_row + 1
        tanggal = row[0].strftime('%d/%m/%Y %H:%M') if row[0] else '-'
        ws.cell(row=r, column=1, value=tanggal)
        ws.cell(row=r, column=2, value=row[1])
        ws.cell(row=r, column=3, value=row[2])
        qty_cell = ws.cell(row=r, column=4, value=row[3])
        qty_cell.alignment = Alignment(horizontal='center')
        if row[3] > 0:
            for c in range(1, 5): ws.cell(row=r, column=c).fill = plus_fill
        else:
            for c in range(1, 5): ws.cell(row=r, column=c).fill = minus_fill

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nama_file = f"stok_{barang_sel.barcode}_{barang_sel.nama[:20].replace(' ','_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nama_file}"'
    wb.save(response)
    return response


# ─── Korporasi: Transfer Stok ──────────────────────────────────────────────────

def transferStokCMS(request):
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    if not cabang.owner_id:
        messages.add_message(request, messages.SUCCESS, "Fitur ini hanya untuk toko Korporasi.")
        return HttpResponseRedirect('/cms/')

    from owner.models import TransferStok
    owner = cabang.owner
    cabang_tujuan_list = owner.cabang_korporasi.exclude(id=cabang.id)

    if request.method == 'POST':
        aksi = request.POST.get('aksi', 'kirim')

        # ── Batalkan transfer pending ──
        if aksi == 'batal':
            transfer_id = request.POST.get('transfer_id')
            try:
                t = TransferStok.objects.get(
                    id=transfer_id,
                    cabang_asal=cabang,
                    status=TransferStok.STATUS_PENDING
                )
                nama_barang = t.barang_asal.nama
                t.delete()
                messages.add_message(request, messages.SUCCESS,
                    f"Permintaan transfer {nama_barang} berhasil dibatalkan.")
            except TransferStok.DoesNotExist:
                messages.add_message(request, messages.SUCCESS,
                    "Transfer tidak ditemukan atau sudah diproses.")
            return HttpResponseRedirect('/cms/transfer/')

        barang_id        = request.POST.get('barang_id')
        cabang_tujuan_id = request.POST.get('cabang_tujuan')
        jumlah           = int(request.POST.get('jumlah', 0))
        catatan          = request.POST.get('catatan', '').strip()

        if jumlah <= 0:
            messages.add_message(request, messages.SUCCESS, "Jumlah transfer harus lebih dari 0.")
            return HttpResponseRedirect('/cms/transfer/')

        try:
            from stock.models import Cabang as CabangModel
            barang_asal   = Barang.objects.get(id=barang_id, cabang=cabang)
            cabang_tujuan = CabangModel.objects.get(id=cabang_tujuan_id, owner=owner)
        except Exception:
            messages.add_message(request, messages.SUCCESS, "Data tidak valid.")
            return HttpResponseRedirect('/cms/transfer/')

        if barang_asal.stok < jumlah:
            messages.add_message(request, messages.SUCCESS,
                f"Stok {barang_asal.nama} hanya {barang_asal.stok}, tidak cukup untuk transfer {jumlah}.")
            return HttpResponseRedirect('/cms/transfer/')

        import string, random as _rnd
        nomor = ''.join(_rnd.choices(string.ascii_uppercase + string.digits, k=10))
        TransferStok.objects.create(
            owner=owner,
            cabang_asal=cabang,
            cabang_tujuan=cabang_tujuan,
            barang_asal=barang_asal,
            jumlah=jumlah,
            catatan=catatan,
            nomor_faktur=nomor,
            status=TransferStok.STATUS_PENDING,
            created_by=request.user,
        )
        messages.add_message(request, messages.SUCCESS,
            f"Permintaan transfer {jumlah} {barang_asal.satuan} {barang_asal.nama} "
            f"→ {cabang_tujuan.nama_toko} berhasil dikirim. Menunggu persetujuan pemilik.")
        return HttpResponseRedirect('/cms/transfer/')

    from owner.models import TransferStok
    barang_list = Barang.objects.filter(cabang=cabang, stok__gt=0).order_by('nama')
    riwayat = TransferStok.objects.filter(cabang_asal=cabang).order_by('-created_at')[:20]

    context = {
        'cabang': cabang,
        'barang_list': barang_list,
        'cabang_tujuan_list': cabang_tujuan_list,
        'riwayat': riwayat,
    }
    # Akses via HTMX → render komponen saja; akses langsung → render halaman penuh
    # Akses langsung (non-HTMX GET) → redirect ke CMS agar sidebar lengkap
    if not request.headers.get('HX-Request') and request.method == 'GET':
        return HttpResponseRedirect('/cms/')
    return render(request, 'administrator/components/transfer_stok.html', context)


# ─── Korporasi: Request Master Barang ke Gudang ────────────────────────────────

def requestMasterBarang(request):
    """Admin toko request penambahan master barang baru ke Gudang Utama."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')
    cabang = request.user.userprofile.cabang
    if not cabang.owner_id:
        return HttpResponseRedirect('/cms/')

    from owner.models import RequestMasterBarang
    owner = cabang.owner

    notif_ok = None
    notif_err = None

    if request.method == 'POST':
        barcode  = request.POST.get('barcode', '').strip()
        nama     = request.POST.get('nama', '').strip()
        satuan   = request.POST.get('satuan', 'PCS')
        catatan  = request.POST.get('catatan_toko', '').strip()
        harga_beli     = int(request.POST.get('harga_beli', 0) or 0)
        harga_ecer     = int(request.POST.get('harga_ecer', 0) or 0)
        harga_grosir   = int(request.POST.get('harga_grosir', 0) or 0)
        min_beli_grosir = int(request.POST.get('min_beli_grosir', 0) or 0)
        keterangan     = request.POST.get('keterangan', '').strip()

        if not barcode or not nama:
            notif_err = "Barcode dan nama barang wajib diisi."
        else:
            RequestMasterBarang.objects.create(
                owner=owner, cabang=cabang,
                barcode=barcode, nama=nama, satuan=satuan,
                harga_beli=harga_beli, harga_ecer=harga_ecer,
                harga_grosir=harga_grosir, min_beli_grosir=min_beli_grosir,
                keterangan=keterangan, catatan_toko=catatan,
                status=RequestMasterBarang.STATUS_PENDING,
                created_by=request.user,
            )
            notif_ok = f"Request master barang '{nama}' berhasil dikirim. Menunggu persetujuan pemilik gudang."

    riwayat = RequestMasterBarang.objects.filter(cabang=cabang).order_by('-created_at')[:20]
    return render(request, 'administrator/components/request_master_barang.html', {
        'cabang': cabang,
        'riwayat': riwayat,
        'notif_ok': notif_ok,
        'notif_err': notif_err,
    })


# ─── Korporasi: Order Barang ke Gudang Utama ──────────────────────────────────

def orderBarang(request):
    """Admin toko mengajukan order barang ke Gudang Utama melalui upload Excel."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    if not cabang.owner_id:
        return HttpResponseRedirect('/cms/')

    from owner.models import OrderBarang, OrderBarangItem
    import string, random as _rnd
    owner = cabang.owner
    notif_ok = None
    notif_err = None

    if request.method == 'POST' and request.FILES.get('file'):
        catatan_toko = request.POST.get('catatan_toko', '').strip()
        nomor = ''.join(_rnd.choices(string.ascii_uppercase + string.digits, k=10))

        try:
            df = pandas.read_excel(request.FILES['file'])
            # Normalisasi nama kolom: lowercase, strip spasi, hapus teks dalam kurung
            import re
            df.columns = [
                re.sub(r'\s*\(.*\)', '', str(c)).strip().lower().replace(' ', '_')
                for c in df.columns
            ]
            gudang = cabang.owner.cabang_korporasi.filter(is_gudang=True).first()
            items_valid = []

            for _, row in df.iterrows():
                try:
                    barcode_raw = row.get('barcode', '')
                    jumlah_raw  = row.get('jumlah', 0)
                    barcode = str(int(float(barcode_raw))) if str(barcode_raw) not in ('nan', '', 'None') else ''
                    jumlah  = int(float(jumlah_raw)) if str(jumlah_raw) not in ('nan', '', 'None', '0.0') else 0
                    if not barcode or jumlah <= 0:
                        continue
                    nama = ''
                    if gudang:
                        try:
                            b = Barang.objects.get(cabang=gudang, barcode=barcode)
                            nama = b.nama
                        except Barang.DoesNotExist:
                            nama = f'Barcode {barcode} (tidak ditemukan di gudang)'
                    items_valid.append({'barcode': barcode, 'nama': nama, 'jumlah': jumlah})
                except Exception:
                    continue

            if not items_valid:
                notif_err = "File tidak valid atau tidak ada baris dengan jumlah > 0 yang dapat dibaca. Pastikan kolom 'barcode' dan 'jumlah' terisi."
            else:
                order = OrderBarang.objects.create(
                    owner=owner, cabang=cabang, nomor_order=nomor,
                    catatan_toko=catatan_toko,
                    status=OrderBarang.STATUS_PENDING,
                    created_by=request.user,
                )
                for item in items_valid:
                    OrderBarangItem.objects.create(
                        order=order, barcode=item['barcode'],
                        nama_barang=item['nama'], jumlah_order=item['jumlah'],
                    )
                notif_ok = f"Order <strong>{nomor}</strong> berhasil dikirim ({len(items_valid)} item). Menunggu persetujuan pemilik gudang."

        except Exception as e:
            notif_err = f"Gagal membaca file Excel: {e}. Pastikan format sesuai template."

    elif request.method == 'POST' and not request.FILES.get('file'):
        notif_err = "Pilih file Excel terlebih dahulu sebelum mengirim order."

    riwayat = OrderBarang.objects.filter(cabang=cabang).order_by('-created_at')[:20]
    context = {
        'cabang': cabang,
        'riwayat': riwayat,
        'notif_ok': notif_ok,
        'notif_err': notif_err,
    }
    return render(request, 'administrator/components/order_barang.html', context)


def downloadTemplateOrder(request):
    """Download template Excel berisi semua barang gudang (barcode, nama, jumlah kosong)."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    if not cabang.owner_id:
        return HttpResponseRedirect('/cms/')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    gudang = cabang.owner.cabang_korporasi.filter(is_gudang=True).first()
    barangs = Barang.objects.filter(cabang=gudang).order_by('nama') if gudang else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Order"

    # Header style
    header_fill = PatternFill(start_color="1E2C5A", end_color="1E2C5A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='D8E3F5'),
        right=Side(style='thin', color='D8E3F5'),
        top=Side(style='thin', color='D8E3F5'),
        bottom=Side(style='thin', color='D8E3F5'),
    )

    headers = ['barcode', 'nama_barang', 'jumlah']
    header_labels = ['Barcode', 'Nama Barang', 'Jumlah (isi di sini)']
    col_widths = [18, 40, 20]

    for col_idx, (key, label, width) in enumerate(zip(headers, header_labels, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # Data rows
    row_fill_alt = PatternFill(start_color="F0F4FB", end_color="F0F4FB", fill_type="solid")
    data_font = Font(size=10)

    for row_idx, b in enumerate(barangs, start=2):
        row_data = [b.barcode, b.nama, '']
        fill = row_fill_alt if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin
            if row_idx % 2 == 0:
                cell.fill = row_fill_alt
            if col_idx == 3:
                cell.alignment = center

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Info sheet
    ws2 = wb.create_sheet("Petunjuk")
    ws2['A1'] = "PETUNJUK PENGISIAN TEMPLATE ORDER BARANG"
    ws2['A1'].font = Font(bold=True, size=12, color="1E2C5A")
    ws2.column_dimensions['A'].width = 60
    petunjuk = [
        "",
        "1. Isi kolom JUMLAH dengan angka jumlah barang yang ingin dipesan.",
        "2. Jangan mengubah kolom Barcode dan Nama Barang.",
        "3. Kosongkan kolom Jumlah jika tidak ingin memesan barang tersebut.",
        "4. Simpan file dan upload kembali di halaman Order Barang.",
        "",
        f"Gudang: {gudang.nama_toko if gudang else '-'}",
        f"Toko: {cabang.nama_toko}",
    ]
    for r, txt in enumerate(petunjuk, start=2):
        ws2.cell(row=r, column=1, value=txt).font = Font(size=10)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nama_file = f"template_order_{cabang.prefix}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nama_file}"'
    wb.save(response)
    return response


# ─── Korporasi: Penerimaan Barang dari Gudang ──────────────────────────────────

def penerimaanBarang(request):
    """Admin toko melihat daftar pengiriman masuk (dari gudang DAN transfer antar toko)."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')

    cabang = request.user.userprofile.cabang
    if not cabang.owner_id:
        return HttpResponseRedirect('/cms/')

    from owner.models import PengirimanGudang, PengirimanGudangItem, TransferStok

    # Pengiriman dari gudang utama
    pengiriman_list = PengirimanGudang.objects.filter(
        cabang_tujuan=cabang
    ).order_by('-created_at')

    # Transfer masuk dari toko lain — pending (menunggu konfirmasi) + history semua
    transfer_masuk_pending = TransferStok.objects.filter(
        cabang_tujuan=cabang,
        status=TransferStok.STATUS_PENDING
    ).order_by('-created_at')

    transfer_masuk_history = TransferStok.objects.filter(
        cabang_tujuan=cabang
    ).exclude(status=TransferStok.STATUS_PENDING).order_by('-updated_at')[:30]

    transfer_masuk = transfer_masuk_pending  # alias untuk template lama

    # Konfirmasi transfer toko-ke-toko dari halaman ini
    if request.method == 'POST':
        transfer_id = request.POST.get('transfer_id')
        aksi        = request.POST.get('aksi_transfer')
        try:
            t = TransferStok.objects.get(id=transfer_id, cabang_tujuan=cabang, status=TransferStok.STATUS_PENDING)
            if aksi == 'terima':
                # Eksekusi: tambah stok tujuan, kurangi stok asal
                try:
                    bt = Barang.objects.get(cabang=cabang, barcode=t.barang_asal.barcode)
                    bt.stok += t.jumlah
                    bt.save()
                except Barang.DoesNotExist:
                    bt = Barang.objects.create(
                        cabang=cabang, barcode=t.barang_asal.barcode,
                        nama=t.barang_asal.nama, satuan=t.barang_asal.satuan,
                        stok=0, harga_beli=t.barang_asal.harga_beli,
                        harga_ecer=t.barang_asal.harga_ecer,
                        harga_grosir=t.barang_asal.harga_grosir,
                        min_beli_grosir=t.barang_asal.min_beli_grosir,
                        keterangan=t.barang_asal.keterangan,
                        created_by=request.user,
                    )
                    bt.stok += t.jumlah
                    bt.save()
                t.barang_asal.stok -= t.jumlah
                t.barang_asal.save()
                t.barang_tujuan = bt
                t.status = TransferStok.STATUS_APPROVED
                t.approved_by = request.user
                t.save()
                messages.add_message(request, messages.SUCCESS,
                    f"{t.barang_asal.nama} diterima. Stok bertambah {t.jumlah}.")
            elif aksi == 'tolak':
                t.status = TransferStok.STATUS_REJECTED
                t.save()
                messages.add_message(request, messages.SUCCESS,
                    f"Transfer {t.barang_asal.nama} ditolak.")
        except TransferStok.DoesNotExist:
            messages.add_message(request, messages.SUCCESS, "Transfer tidak ditemukan.")

        # Jika via HTMX → render ulang komponen; jika POST langsung → redirect CMS
        if request.headers.get('HX-Request'):
            pengiriman_list_updated = PengirimanGudang.objects.filter(cabang_tujuan=cabang).order_by('-created_at')
            tm_pending  = TransferStok.objects.filter(cabang_tujuan=cabang, status=TransferStok.STATUS_PENDING).order_by('-created_at')
            tm_history  = TransferStok.objects.filter(cabang_tujuan=cabang).exclude(status=TransferStok.STATUS_PENDING).order_by('-updated_at')[:30]
            return render(request, 'administrator/components/penerimaan_barang.html', {
                'cabang': cabang,
                'pengiriman_list': pengiriman_list_updated,
                'transfer_masuk': tm_pending,
                'transfer_masuk_pending': tm_pending,
                'transfer_masuk_history': tm_history,
            })
        return HttpResponseRedirect('/cms/')

    # Akses langsung (non-HTMX) → redirect ke CMS agar sidebar muncul
    if not request.headers.get('HX-Request'):
        return HttpResponseRedirect('/cms/')

    return render(request, 'administrator/components/penerimaan_barang.html', {
        'cabang': cabang,
        'pengiriman_list': pengiriman_list,
        'transfer_masuk': transfer_masuk,
        'transfer_masuk_pending': transfer_masuk_pending,
        'transfer_masuk_history': transfer_masuk_history,
    })


def konfirmasiItemPenerimaan(request, pengiriman_id):
    """Admin toko konfirmasi per item: diterima atau dikembalikan."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')
    # Akses langsung → tetap render halaman (tidak redirect)

    cabang = request.user.userprofile.cabang
    from owner.models import PengirimanGudang, PengirimanGudangItem

    try:
        pengiriman = PengirimanGudang.objects.get(id=pengiriman_id, cabang_tujuan=cabang)
    except PengirimanGudang.DoesNotExist:
        messages.add_message(request, messages.SUCCESS, "Pengiriman tidak ditemukan.")
        return HttpResponseRedirect('/cms/penerimaan/')

    if request.method == 'POST':
        item_id      = request.POST.get('item_id')
        aksi         = request.POST.get('aksi')
        catatan_toko = request.POST.get('catatan_toko', '').strip()

        try:
            item = PengirimanGudangItem.objects.get(
                id=item_id, pengiriman=pengiriman,
                status=PengirimanGudangItem.STATUS_PENDING
            )
        except PengirimanGudangItem.DoesNotExist:
            messages.add_message(request, messages.SUCCESS, "Item tidak valid atau sudah dikonfirmasi.")
            return HttpResponseRedirect(f'/cms/penerimaan/{pengiriman_id}/')

        if aksi == 'terima':
            # Cari atau buat barang di toko, tambah stok
            try:
                bt = Barang.objects.get(cabang=cabang, barcode=item.barang_gudang.barcode)
                bt.stok += item.jumlah_dikirim
                bt.save()
            except Barang.DoesNotExist:
                Barang.objects.create(
                    cabang=cabang,
                    barcode=item.barang_gudang.barcode,
                    nama=item.barang_gudang.nama,
                    satuan=item.barang_gudang.satuan,
                    stok=item.jumlah_dikirim,
                    harga_beli=item.barang_gudang.harga_beli,
                    harga_ecer=item.barang_gudang.harga_ecer,
                    harga_grosir=item.barang_gudang.harga_grosir,
                    min_beli_grosir=item.barang_gudang.min_beli_grosir,
                    keterangan=item.barang_gudang.keterangan,
                    created_by=request.user,
                )
            item.jumlah_diterima = item.jumlah_dikirim
            item.status      = PengirimanGudangItem.STATUS_DITERIMA
            item.catatan_toko = catatan_toko
            item.save()
            messages.add_message(request, messages.SUCCESS, f"{item.barang_gudang.nama} diterima. Stok toko bertambah {item.jumlah_dikirim}.")

        elif aksi == 'kembalikan':
            item.status      = PengirimanGudangItem.STATUS_DIKEMBALIKAN
            item.catatan_toko = catatan_toko
            item.jumlah_diterima = 0
            item.save()
            messages.add_message(request, messages.SUCCESS, f"{item.barang_gudang.nama} dikembalikan ke gudang. Menunggu konfirmasi pemilik.")

        # Update status pengiriman — selesai jika tidak ada yang pending
        items = pengiriman.item_set.all()
        if not items.filter(status=PengirimanGudangItem.STATUS_PENDING).exists():
            pengiriman.status = PengirimanGudang.STATUS_SELESAI
            pengiriman.save()

        # Setelah POST → PRG redirect ke halaman yang sama
        return HttpResponseRedirect(f'/cms/penerimaan/{pengiriman_id}/')

    items = pengiriman.item_set.all()
    context = {
        'cabang': cabang,
        'pengiriman': pengiriman,
        'items': items,
    }
    return render(request, 'administrator/penerimaan_konfirmasi.html', context)


# ─── Korporasi: Daftar Barang (read-only + download) ───────────────────────────

def daftarBarangKorporasi(request):
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponseRedirect('/login/')
    cabang = request.user.userprofile.cabang
    barangs = Barang.objects.filter(cabang=cabang).order_by('nama')
    return render(request, 'administrator/components/daftar_barang_korporasi.html', {
        'barangs': barangs,
        'jumlah_barang': barangs.count(),
    })


# ─── Master Data Views (superuser only) ───────────────────────────────────────

def _superuser_required(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect('/login/')
    if not request.user.is_superuser:
        messages.add_message(request, messages.ERROR, "Halaman ini hanya untuk superadmin.")
        return HttpResponseRedirect('/cms/')
    return None


def masterPaket(request):
    redirect = _superuser_required(request)
    if redirect:
        return redirect
    if request.method == 'POST':
        aksi = request.POST.get('aksi', '')
        if aksi == 'tambah':
            paket = DaftarPaket()
            paket.nama = request.POST.get('nama', '')
            paket.max_transaksi = int(request.POST.get('max_transaksi', 0))
            paket.max_user_login = int(request.POST.get('max_user_login', 0))
            paket.harga_per_bulan = int(request.POST.get('harga_per_bulan', 0))
            paket.harga_per_tiga_bulan = int(request.POST.get('harga_per_tiga_bulan', 0))
            paket.harga_per_enam_bulan = int(request.POST.get('harga_per_enam_bulan', 0))
            paket.harga_per_tahun = int(request.POST.get('harga_per_tahun', 0))
            paket.harga_per_dua_tahun = int(request.POST.get('harga_per_dua_tahun', 0))
            paket.disc = int(request.POST.get('disc', 0))
            paket.save()
            messages.success(request, f"Paket '{paket.nama}' berhasil ditambahkan.")
        elif aksi == 'edit':
            paket_id = request.POST.get('paket_id')
            paket = get_object_or_404(DaftarPaket, pk=paket_id)
            paket.nama = request.POST.get('nama', paket.nama)
            paket.max_transaksi = int(request.POST.get('max_transaksi', paket.max_transaksi))
            paket.max_user_login = int(request.POST.get('max_user_login', paket.max_user_login))
            paket.harga_per_bulan = int(request.POST.get('harga_per_bulan', paket.harga_per_bulan))
            paket.harga_per_tiga_bulan = int(request.POST.get('harga_per_tiga_bulan', paket.harga_per_tiga_bulan))
            paket.harga_per_enam_bulan = int(request.POST.get('harga_per_enam_bulan', paket.harga_per_enam_bulan))
            paket.harga_per_tahun = int(request.POST.get('harga_per_tahun', paket.harga_per_tahun))
            paket.harga_per_dua_tahun = int(request.POST.get('harga_per_dua_tahun', paket.harga_per_dua_tahun))
            paket.disc = int(request.POST.get('disc', paket.disc))
            paket.save()
            messages.success(request, f"Paket '{paket.nama}' berhasil diperbarui.")
        elif aksi == 'hapus':
            paket_id = request.POST.get('paket_id')
            paket = get_object_or_404(DaftarPaket, pk=paket_id)
            nama = paket.nama
            paket.delete()
            messages.success(request, f"Paket '{nama}' berhasil dihapus.")
    pakets = DaftarPaket.objects.all()
    return render(request, 'administrator/components/master_paket.html', {'pakets': pakets})


def masterPromo(request):
    redirect = _superuser_required(request)
    if redirect:
        return redirect
    if request.method == 'POST':
        aksi = request.POST.get('aksi', '')
        if aksi == 'tambah':
            promo = Promo()
            promo.nama = request.POST.get('nama', '')
            promo.kode = request.POST.get('kode', '').upper().strip()
            promo.disc = int(request.POST.get('disc', 0))
            promo.kuota = int(request.POST.get('kuota', 0))
            end_period = request.POST.get('end_period', '')
            promo.end_period = end_period if end_period else None
            promo.is_registrasi = request.POST.get('is_registrasi') == 'on'
            promo.is_perpanjangan = request.POST.get('is_perpanjangan') == 'on'
            promo.is_upgrade_lisensi = request.POST.get('is_upgrade_lisensi') == 'on'
            promo.is_active = True
            promo.save()
            messages.success(request, f"Promo '{promo.nama}' berhasil ditambahkan.")
        elif aksi == 'edit':
            promo_id = request.POST.get('promo_id')
            promo = get_object_or_404(Promo, pk=promo_id)
            promo.nama = request.POST.get('nama', promo.nama)
            promo.kode = request.POST.get('kode', promo.kode).upper().strip()
            promo.disc = int(request.POST.get('disc', promo.disc))
            promo.kuota = int(request.POST.get('kuota', promo.kuota))
            end_period = request.POST.get('end_period', '')
            promo.end_period = end_period if end_period else None
            promo.is_registrasi = request.POST.get('is_registrasi') == 'on'
            promo.is_perpanjangan = request.POST.get('is_perpanjangan') == 'on'
            promo.is_upgrade_lisensi = request.POST.get('is_upgrade_lisensi') == 'on'
            promo.save()
            messages.success(request, f"Promo '{promo.nama}' berhasil diperbarui.")
        elif aksi == 'hapus':
            promo_id = request.POST.get('promo_id')
            promo = get_object_or_404(Promo, pk=promo_id)
            nama = promo.nama
            promo.delete()
            messages.success(request, f"Promo '{nama}' berhasil dihapus.")
        elif aksi == 'toggle':
            promo_id = request.POST.get('promo_id')
            promo = get_object_or_404(Promo, pk=promo_id)
            promo.is_active = not promo.is_active
            promo.save()
            status = "diaktifkan" if promo.is_active else "dinonaktifkan"
            messages.success(request, f"Promo '{promo.nama}' berhasil {status}.")
    promos = Promo.objects.all().order_by('-created_at')
    return render(request, 'administrator/components/master_promo.html', {'promos': promos})


def masterTestimoni(request):
    redirect = _superuser_required(request)
    if redirect:
        return redirect
    if request.method == 'POST':
        aksi = request.POST.get('aksi', '')
        testimoni_id = request.POST.get('testimoni_id')
        testimoni = get_object_or_404(Testimoni, pk=testimoni_id)
        if aksi == 'approve':
            testimoni.is_verified = True
            testimoni.save()
            messages.success(request, "Testimoni berhasil disetujui.")
        elif aksi == 'reject':
            testimoni.is_verified = False
            testimoni.save()
            messages.success(request, "Testimoni ditolak.")
        elif aksi == 'hapus':
            testimoni.delete()
            messages.success(request, "Testimoni berhasil dihapus.")
    testimonis = Testimoni.objects.all().order_by('-created_at')
    return render(request, 'administrator/components/master_testimoni.html', {'testimonis': testimonis})


def masterCabang(request):
    redirect = _superuser_required(request)
    if redirect:
        return redirect
    if request.method == 'POST':
        aksi = request.POST.get('aksi', '')
        cabang_id = request.POST.get('cabang_id')
        cabang = get_object_or_404(Cabang, pk=cabang_id)
        if aksi == 'toggle':
            cabang.is_active = not cabang.is_active
            cabang.save()
            status = "diaktifkan" if cabang.is_active else "dinonaktifkan"
            messages.success(request, f"Cabang '{cabang.nama_toko}' berhasil {status}.")
        elif aksi == 'set_paket':
            paket_id = request.POST.get('paket_id')
            if paket_id:
                paket = get_object_or_404(DaftarPaket, pk=paket_id)
                cabang.paket = paket
                cabang.kuota_transaksi = paket.max_transaksi
            else:
                cabang.paket = None
                cabang.lisensi_expired = None
                cabang.lisensi_grace = None
                cabang.kuota_transaksi = 75
            cabang.save()
            messages.success(request, f"Paket cabang '{cabang.nama_toko}' berhasil diperbarui.")
    cabangs = Cabang.objects.all().order_by('-created_at')
    pakets = DaftarPaket.objects.all()
    return render(request, 'administrator/components/master_cabang.html', {'cabangs': cabangs, 'pakets': pakets})


# ─── Tempo Payment ──────────────────────────────────────────────────────────────

def _cek_akses_tempo(request):
    """Return (cabang, error_response) — error_response is None jika akses OK."""
    if not request.user.is_authenticated:
        return None, HttpResponseRedirect('/login/')
    if not request.user.is_superuser:
        messages.add_message(request, messages.SUCCESS, 'Hanya pemilik toko yang dapat mengakses halaman ini.')
        return None, HttpResponseRedirect('/')
    cabang = request.user.userprofile.cabang
    if cabang.paket is None and cabang.owner_id is None:
        messages.add_message(request, messages.SUCCESS, 'Fitur Tempo hanya tersedia untuk paket berbayar.')
        return None, HttpResponseRedirect('/cms/')
    return cabang, None


def daftarTempoTransaksi(request):
    cabang, err = _cek_akses_tempo(request)
    if err:
        return err

    now = datetime.datetime.now()

    qs = Penjualan.objects.filter(
        cabang=cabang, metode=2, is_paid=True, is_tempo_lunas=False, is_void=False
    ).select_related('user__userprofile').order_by('jatuh_tempo')

    bulan_berjalan = [t for t in qs if t.created_at.month == now.month and t.created_at.year == now.year]
    bulan_sebelumnya = [t for t in qs if not (t.created_at.month == now.month and t.created_at.year == now.year)]

    BULAN_LIST = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']

    context = {
        'bulan_berjalan': bulan_berjalan,
        'bulan_sebelumnya': bulan_sebelumnya,
        'bulan_ini': BULAN_LIST[now.month],
        'tahun_ini': now.year,
        'total_tempo': qs.count(),
    }
    return render(request, 'administrator/components/tempo_list.html', context)


def bayarTempoTransaksi(request):
    cabang, err = _cek_akses_tempo(request)
    if err:
        return err

    nota = request.GET.get('id') or request.POST.get('nota')
    if not nota:
        messages.add_message(request, messages.SUCCESS, 'ID transaksi tidak ditemukan.')
        return HttpResponseRedirect('/cms/tempo/')

    try:
        penjualan = Penjualan.objects.get(
            nota=nota, cabang=cabang, metode=2, is_paid=True, is_tempo_lunas=False
        )
    except Penjualan.DoesNotExist:
        messages.add_message(request, messages.SUCCESS, 'Transaksi tempo tidak ditemukan atau sudah lunas.')
        return HttpResponseRedirect('/cms/tempo/')

    penjualan_detail = PenjualanDetail.objects.filter(penjualan=penjualan).select_related('barang')

    if request.method == 'POST':
        try:
            metode_lunas = int(request.POST.get('metode_lunas', 0))
            customer = request.POST.get('pembeli', penjualan.customer or '')
            returned_items = []

            with __import__('django.db', fromlist=['transaction']).transaction.atomic():
                for detail in penjualan_detail:
                    key = f'jumlah_{detail.id}'
                    try:
                        jumlah_baru = int(request.POST.get(key, detail.jumlah))
                    except (ValueError, TypeError):
                        jumlah_baru = detail.jumlah

                    jumlah_baru = max(0, min(jumlah_baru, detail.jumlah))
                    delta_kembali = detail.jumlah - jumlah_baru

                    if delta_kembali > 0:
                        returned_items.append({
                            'nama': detail.barang.nama,
                            'jumlah_kembali': delta_kembali,
                            'harga': detail.harga,
                        })
                        barang = detail.barang
                        barang.stok += delta_kembali
                        barang.jumlah_dibeli -= delta_kembali
                        barang.save(update_fields=['stok', 'jumlah_dibeli'])

                    if jumlah_baru == 0:
                        detail.delete()
                    else:
                        detail.jumlah = jumlah_baru
                        detail.save()

                # Recalculate totals
                from django.db.models import Sum as _Sum
                agg = PenjualanDetail.objects.filter(penjualan=penjualan).aggregate(
                    total=_Sum('total'), items=Count('id')
                )
                penjualan.total = agg['total'] or 0
                penjualan.items = agg['items'] or 0
                penjualan.is_tempo_lunas = True
                penjualan.metode = metode_lunas
                penjualan.tgl_bayar = datetime.datetime.now()
                penjualan.customer = customer
                penjualan.save()

            addLog(request.user, cabang, 'tempo_lunas',
                   f'Pembayaran tempo lunas nota: {penjualan.nota}')

            # Redirect ke print receipt
            return HttpResponseRedirect(f'/cms/tempo/print/?id={penjualan.nota}')

        except Exception as ex:
            messages.add_message(request, messages.SUCCESS, f'Terjadi kesalahan: {ex}')
            return HttpResponseRedirect(f'/cms/tempo/bayar/?id={nota}')

    context = {
        'penjualan': penjualan,
        'penjualan_detail': penjualan_detail,
    }
    return render(request, 'administrator/components/tempo_bayar.html', context)


def printLunasTempo(request):
    cabang, err = _cek_akses_tempo(request)
    if err:
        return err

    nota = request.GET.get('id', '')
    try:
        penjualan = Penjualan.objects.get(
            nota=nota, cabang=cabang, metode__in=[0, 1], is_tempo_lunas=True
        )
        penjualan_detail = PenjualanDetail.objects.filter(penjualan=penjualan).select_related('barang')
        nama_toko = cabang.nama_toko
        alamat_toko = cabang.alamat_toko
        telpon_toko = cabang.telpon
        total = int(penjualan.total)
        context = {
            'penjualan': penjualan,
            'penjualan_detail': penjualan_detail,
            'nama_toko': nama_toko,
            'alamat_toko': alamat_toko,
            'telpon_toko': telpon_toko,
            'total': total,
        }
        return render(request, 'pos/print_lunas_tempo.html', context)
    except Exception:
        messages.add_message(request, messages.SUCCESS, 'Nota tidak ditemukan.')
        return HttpResponseRedirect('/cms/tempo/')
